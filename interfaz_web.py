"""
=============================================================
🤖 AGENTE PRO ULTIMATE - Interfaz Web con Streamlit
=============================================================

INSTALACIÓN:
  pip install streamlit

EJECUCIÓN:
  streamlit run interfaz_web.py

Se abrirá automáticamente en tu navegador en http://localhost:8501
=============================================================
"""

import streamlit as st
import os
import json
import sqlite3
import smtplib
import csv
import time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from anthropic import Anthropic

load_dotenv()


# =============================================================
# CONFIGURACIÓN DE LA PÁGINA
# =============================================================

st.set_page_config(
    page_title="Agente Pro Ultimate",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado para una interfaz bonita
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700&family=JetBrains+Mono:wght@400;500&display=swap');

    .main { font-family: 'DM Sans', sans-serif; }

    .stApp {
        background: linear-gradient(135deg, #0f0f23 0%, #1a1a3e 50%, #0d0d2b 100%);
    }

    /* Chat messages */
    .user-msg {
        background: linear-gradient(135deg, #2563eb, #1d4ed8);
        color: white;
        padding: 14px 20px;
        border-radius: 18px 18px 4px 18px;
        margin: 8px 0;
        max-width: 85%;
        margin-left: auto;
        font-size: 15px;
        line-height: 1.6;
        box-shadow: 0 2px 12px rgba(37, 99, 235, 0.3);
    }

    .agent-msg {
        background: rgba(255, 255, 255, 0.06);
        color: #e2e8f0;
        padding: 16px 20px;
        border-radius: 18px 18px 18px 4px;
        margin: 8px 0;
        max-width: 85%;
        font-size: 15px;
        line-height: 1.7;
        border: 1px solid rgba(255, 255, 255, 0.08);
        backdrop-filter: blur(10px);
    }

    .tool-badge {
        display: inline-flex;
        align-items: center;
        gap: 6px;
        background: rgba(16, 185, 129, 0.15);
        color: #34d399;
        padding: 4px 12px;
        border-radius: 20px;
        font-size: 12px;
        font-family: 'JetBrains Mono', monospace;
        margin: 4px 2px;
        border: 1px solid rgba(16, 185, 129, 0.2);
    }

    /* Sidebar */
    [data-testid="stSidebar"] {
        background: rgba(15, 15, 35, 0.95) !important;
        border-right: 1px solid rgba(255, 255, 255, 0.06);
    }

    .sidebar-title {
        font-size: 24px;
        font-weight: 700;
        background: linear-gradient(135deg, #60a5fa, #a78bfa);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 4px;
    }

    .sidebar-subtitle {
        font-size: 13px;
        color: #94a3b8;
        margin-bottom: 20px;
    }

    .tool-card {
        background: rgba(255, 255, 255, 0.04);
        border: 1px solid rgba(255, 255, 255, 0.06);
        border-radius: 10px;
        padding: 10px 14px;
        margin: 6px 0;
        display: flex;
        align-items: center;
        gap: 10px;
        transition: all 0.2s;
    }

    .tool-card:hover {
        background: rgba(255, 255, 255, 0.08);
        border-color: rgba(255, 255, 255, 0.12);
    }

    .tool-icon { font-size: 18px; }
    .tool-name { font-size: 13px; color: #e2e8f0; font-weight: 500; }
    .tool-status { font-size: 11px; margin-left: auto; }
    .status-on { color: #34d399; }
    .status-off { color: #fbbf24; }

    /* Header */
    .header-container {
        text-align: center;
        padding: 30px 0 20px;
    }

    .header-icon { font-size: 48px; margin-bottom: 8px; }

    .header-title {
        font-size: 32px;
        font-weight: 700;
        background: linear-gradient(135deg, #60a5fa, #a78bfa, #f472b6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }

    .header-desc {
        font-size: 15px;
        color: #94a3b8;
        margin-top: 4px;
    }

    /* Stats */
    .stats-row {
        display: flex;
        gap: 12px;
        justify-content: center;
        margin: 16px 0 24px;
    }

    .stat-card {
        background: rgba(255, 255, 255, 0.04);
        border: 1px solid rgba(255, 255, 255, 0.06);
        border-radius: 12px;
        padding: 12px 20px;
        text-align: center;
        min-width: 120px;
    }

    .stat-number {
        font-size: 24px;
        font-weight: 700;
        color: #60a5fa;
    }

    .stat-label {
        font-size: 12px;
        color: #64748b;
        margin-top: 2px;
    }

    /* Input */
    .stTextInput > div > div > input {
        background: rgba(255, 255, 255, 0.06) !important;
        border: 1px solid rgba(255, 255, 255, 0.1) !important;
        border-radius: 14px !important;
        color: #e2e8f0 !important;
        padding: 14px 18px !important;
        font-size: 15px !important;
        font-family: 'DM Sans', sans-serif !important;
    }

    .stTextInput > div > div > input:focus {
        border-color: rgba(96, 165, 250, 0.5) !important;
        box-shadow: 0 0 0 3px rgba(96, 165, 250, 0.1) !important;
    }

    .stTextInput > div > div > input::placeholder {
        color: #475569 !important;
    }

    /* Quick actions */
    .quick-action {
        background: rgba(255, 255, 255, 0.04);
        border: 1px solid rgba(255, 255, 255, 0.08);
        border-radius: 10px;
        padding: 10px 16px;
        color: #94a3b8;
        font-size: 13px;
        cursor: pointer;
        transition: all 0.2s;
        text-align: left;
    }

    .quick-action:hover {
        background: rgba(255, 255, 255, 0.08);
        color: #e2e8f0;
        border-color: rgba(96, 165, 250, 0.3);
    }

    /* Hide Streamlit elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}

    div[data-testid="stDecoration"] { display: none; }
</style>
""", unsafe_allow_html=True)


# =============================================================
# TODAS LAS HERRAMIENTAS (mismas del agente_pro_ultimate.py)
# =============================================================

def buscar_en_web(consulta, max_resultados=5):
    try:
        from duckduckgo_search import DDGS
    except ImportError:
        return "Error: Instala duckduckgo-search"
    for intento in range(3):
        try:
            with DDGS() as ddgs:
                resultados = list(ddgs.text(consulta, max_results=max_resultados, region="wt-wt"))
            if resultados:
                texto = f"Resultados para '{consulta}':\n\n"
                for i, r in enumerate(resultados, 1):
                    texto += f"{i}. **{r['title']}**\n   {r['body']}\n   Fuente: {r['href']}\n\n"
                return texto
        except Exception:
            if intento < 2: time.sleep(2)
    try:
        with DDGS() as ddgs:
            resultados = list(ddgs.news(consulta, max_results=max_resultados))
        if resultados:
            texto = f"Noticias para '{consulta}':\n\n"
            for i, r in enumerate(resultados, 1):
                texto += f"{i}. **{r['title']}**\n   {r['body']}\n   Fuente: {r['source']}\n\n"
            return texto
    except Exception:
        pass
    return f"No se encontraron resultados para '{consulta}'."

def enviar_correo(destinatario, asunto, cuerpo):
    ea, ep = os.getenv("EMAIL_ADDRESS"), os.getenv("EMAIL_PASSWORD")
    if not ea or not ep:
        return "Error: Configura EMAIL_ADDRESS y EMAIL_PASSWORD en .env"
    try:
        msg = MIMEMultipart()
        msg["From"], msg["To"], msg["Subject"] = ea, destinatario, asunto
        msg.attach(MIMEText(cuerpo, "plain"))
        with smtplib.SMTP("smtp.gmail.com", 587) as s:
            s.starttls(); s.login(ea, ep); s.send_message(msg)
        return f"Correo enviado a {destinatario}"
    except Exception as e:
        return f"Error: {str(e)}"

def leer_pdf(ruta_archivo):
    try:
        from PyPDF2 import PdfReader
    except ImportError:
        return "Error: Instala PyPDF2"
    try:
        ruta = Path(ruta_archivo)
        if not ruta.exists(): return f"No se encontró '{ruta_archivo}'"
        reader = PdfReader(str(ruta))
        texto = ""
        for i, p in enumerate(reader.pages, 1):
            t = p.extract_text()
            if t: texto += f"\n--- Página {i} ---\n{t}"
        if not texto.strip(): return "PDF sin texto extraíble."
        if len(texto) > 4000: texto = texto[:4000] + "\n[... truncado ...]"
        return f"PDF '{ruta.name}' ({len(reader.pages)} págs):\n{texto}"
    except Exception as e:
        return f"Error: {str(e)}"

def leer_excel_csv(ruta_archivo, max_filas=20):
    try:
        ruta = Path(ruta_archivo)
        if not ruta.exists(): return f"No se encontró '{ruta_archivo}'"
        ext = ruta.suffix.lower()
        if ext == ".csv":
            filas = []
            with open(ruta, "r", encoding="utf-8-sig") as f:
                for i, fila in enumerate(csv.reader(f)):
                    if i >= max_filas + 1: break
                    filas.append(fila)
        elif ext in [".xlsx", ".xls"]:
            from openpyxl import load_workbook
            wb = load_workbook(str(ruta), read_only=True); ws = wb.active
            filas = [[str(c) if c else "" for c in r] for i, r in enumerate(ws.iter_rows(values_only=True)) if i < max_filas + 1]
            wb.close()
        else:
            return f"Formato no soportado: {ext}"
        if not filas: return "Archivo vacío."
        enc, dat = filas[0], filas[1:]
        res = f"Archivo: {ruta.name}\nColumnas: {', '.join(enc)}\n\n"
        res += " | ".join(enc) + "\n" + "-" * 40 + "\n"
        for f in dat: res += " | ".join(f) + "\n"
        return res
    except Exception as e:
        return f"Error: {str(e)}"

DB_PATH = "agente_datos.db"

def inicializar_db():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS contactos (id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL, email TEXT, telefono TEXT, notas TEXT, creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    c.execute("CREATE TABLE IF NOT EXISTS tareas (id INTEGER PRIMARY KEY AUTOINCREMENT, titulo TEXT NOT NULL, descripcion TEXT, estado TEXT DEFAULT 'pendiente', prioridad TEXT DEFAULT 'media', creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    c.execute("SELECT COUNT(*) FROM contactos")
    if c.fetchone()[0] == 0:
        c.executemany("INSERT INTO contactos (nombre, email, telefono, notas) VALUES (?, ?, ?, ?)",
            [("Ana García","ana@ejemplo.com","33-1234-5678","Cliente principal"),
             ("Carlos López","carlos@ejemplo.com","81-9876-5432","Proveedor"),
             ("María Rodríguez","maria@ejemplo.com","55-5555-1234","Socia")])
    c.execute("SELECT COUNT(*) FROM tareas")
    if c.fetchone()[0] == 0:
        c.executemany("INSERT INTO tareas (titulo, descripcion, estado, prioridad) VALUES (?, ?, ?, ?)",
            [("Revisar propuesta","Propuesta del cliente","pendiente","alta"),
             ("Enviar reporte","Reporte mensual","pendiente","media"),
             ("Llamar a proveedor","Negociar precios","completada","baja")])
    conn.commit(); conn.close()

def consultar_base_datos(sql):
    up = sql.strip().upper()
    if any(p in up for p in ["DROP","DELETE","ALTER","TRUNCATE"]): return "Operación no permitida."
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor(); c.execute(sql)
        if up.startswith("SELECT"):
            cols = [d[0] for d in c.description] if c.description else []
            filas = c.fetchall()
            if not filas: res = "Sin resultados."
            else:
                res = f"Columnas: {', '.join(cols)}\n\n"
                for f in filas: res += " | ".join(str(v) for v in f) + "\n"
        else:
            conn.commit(); res = f"Ejecutado. Filas: {c.rowcount}"
        conn.close(); return res
    except Exception as e:
        return f"Error: {str(e)}"

def guardar_archivo(nombre, contenido):
    try:
        r = Path(nombre)
        if r.suffix.lower() not in [".txt",".md",".csv",".json",".html"]: return "Extensión no permitida."
        with open(r, "w", encoding="utf-8") as f: f.write(contenido)
        return f"Archivo '{nombre}' guardado ({len(contenido)} chars)."
    except Exception as e:
        return f"Error: {str(e)}"

def traducir_texto(texto, idioma_destino, idioma_origen="auto"):
    try:
        client = Anthropic()
        resp = client.messages.create(model="claude-sonnet-4-6", max_tokens=2048,
            messages=[{"role":"user","content":f"Traduce al {idioma_destino}. Solo la traducción:\n{texto}"}])
        trad = "".join(b.text for b in resp.content if hasattr(b,"text"))
        return f"Traducción al {idioma_destino}:\n\n{trad}"
    except Exception as e:
        return f"Error: {str(e)}"

def generar_grafica(ruta_csv, tipo="barras", col_x="", col_y="", titulo="Gráfica", archivo="grafica.png"):
    try:
        import matplotlib; matplotlib.use("Agg")
        import matplotlib.pyplot as plt; import matplotlib.ticker as ticker
    except ImportError:
        return "Error: Instala matplotlib"
    try:
        ruta = Path(ruta_csv)
        if not ruta.exists(): return f"No se encontró '{ruta_csv}'"
        datos, enc = [], []
        if ruta.suffix.lower() == ".csv":
            with open(ruta,"r",encoding="utf-8-sig") as f:
                r = csv.reader(f); enc = next(r); datos = list(r)
        elif ruta.suffix.lower() in [".xlsx",".xls"]:
            from openpyxl import load_workbook
            wb = load_workbook(str(ruta),read_only=True); ws = wb.active
            filas = list(ws.iter_rows(values_only=True)); wb.close()
            enc = [str(c) if c else f"C{i}" for i,c in enumerate(filas[0])]
            datos = [[str(c) if c else "" for c in f] for f in filas[1:]]
        if not col_x: col_x = enc[0]
        if not col_y:
            for i,h in enumerate(enc[1:],1):
                try: float(datos[0][i].replace(",","").replace("$","")); col_y=h; break
                except: continue
            if not col_y: col_y = enc[1] if len(enc)>1 else enc[0]
        xi = yi = None
        for i,h in enumerate(enc):
            if h.strip().lower()==col_x.strip().lower(): xi=i
            if h.strip().lower()==col_y.strip().lower(): yi=i
        if xi is None: return f"Columna '{col_x}' no encontrada. Disponibles: {', '.join(enc)}"
        if yi is None: return f"Columna '{col_y}' no encontrada. Disponibles: {', '.join(enc)}"
        ag = {}
        for f in datos:
            try:
                xv = f[xi].strip()
                yv = float(f[yi].strip().replace(",","").replace("$",""))
                ag[xv] = ag.get(xv,0)+yv
            except: continue
        if not ag: return "Sin datos numéricos."
        ex, vy = list(ag.keys()), list(ag.values())
        plt.style.use("seaborn-v0_8-whitegrid"); fig, ax = plt.subplots(figsize=(10,6))
        cols = ["#378ADD","#1D9E75","#D85A30","#D4537E","#BA7517","#534AB7","#639922","#E24B4A"]
        t = tipo.lower()
        if t in ["barras","bar"]:
            bs = ax.bar(ex,vy,color=cols[:len(ex)],edgecolor="white")
            for b,v in zip(bs,vy): ax.text(b.get_x()+b.get_width()/2.,b.get_height(),f'{v:,.0f}',ha='center',va='bottom',fontsize=9,fontweight='bold')
        elif t in ["lineas","linea","line"]:
            ax.plot(ex,vy,color=cols[0],marker="o",linewidth=2,markersize=8,markerfacecolor="white",markeredgecolor=cols[0],markeredgewidth=2)
        elif t in ["pastel","pie"]:
            ax.pie(vy,labels=ex,colors=cols[:len(ex)],autopct='%1.1f%%',startangle=90); ax.axis('equal')
        elif t in ["barras_horizontales","horizontal","barh"]:
            ax.barh(ex,vy,color=cols[:len(ex)],edgecolor="white")
        else:
            ax.scatter(range(len(vy)),vy,c=cols[0],s=100,edgecolors="white",zorder=5)
            ax.set_xticks(range(len(ex))); ax.set_xticklabels(ex)
        ax.set_title(titulo,fontsize=16,fontweight="bold",pad=15)
        if t not in ["pastel","pie"]:
            ax.set_xlabel(col_x,fontsize=12); ax.set_ylabel(col_y,fontsize=12)
            plt.xticks(rotation=45 if len(ex)>5 else 0,ha='right')
        plt.tight_layout()
        if Path(archivo).suffix.lower() not in [".png",".jpg",".pdf",".svg"]: archivo+=".png"
        plt.savefig(archivo,dpi=150,bbox_inches='tight',facecolor='white'); plt.close()
        return f"Gráfica guardada en: {archivo}"
    except Exception as e:
        return f"Error: {str(e)}"

def resumir_texto(texto, longitud="medio", idioma="español"):
    try:
        inst = {"corto":"3 oraciones máximo","medio":"1-2 párrafos","largo":"detallado por temas","puntos":"5-7 viñetas"}
        client = Anthropic()
        resp = client.messages.create(model="claude-sonnet-4-6", max_tokens=2048,
            messages=[{"role":"user","content":f"Resume en {idioma}. Formato: {inst.get(longitud,'1-2 párrafos')}.\n\nTexto:\n{texto}"}])
        return "".join(b.text for b in resp.content if hasattr(b,"text"))
    except Exception as e:
        return f"Error: {str(e)}"

def obtener_clima(ciudad, pais="MX"):
    ak = os.getenv("OPENWEATHER_API_KEY")
    if not ak: return f"Configura OPENWEATHER_API_KEY en .env (gratis en openweathermap.org)"
    try:
        import requests
        r = requests.get("https://api.openweathermap.org/data/2.5/weather",
            params={"q":f"{ciudad},{pais}","appid":ak,"units":"metric","lang":"es"},timeout=10)
        d = r.json()
        if r.status_code!=200: return f"Error: {d.get('message','Ciudad no encontrada')}"
        return (f"Clima en {d['name']}, {d['sys']['country']}:\n"
                f"🌡️ {d['main']['temp']:.1f}°C (sensación {d['main']['feels_like']:.1f}°C)\n"
                f"📊 Mín: {d['main']['temp_min']:.1f}°C | Máx: {d['main']['temp_max']:.1f}°C\n"
                f"🌤️ {d['weather'][0]['description'].capitalize()}\n"
                f"💧 Humedad: {d['main']['humidity']}% | 💨 Viento: {d['wind']['speed']} m/s")
    except Exception as e:
        return f"Error: {str(e)}"


# Herramientas y mapeo (mismo formato que el agente de terminal)
HERRAMIENTAS = [
    {"name":"buscar_en_web","description":"Busca información actualizada en internet.","input_schema":{"type":"object","properties":{"consulta":{"type":"string","description":"Búsqueda"},"max_resultados":{"type":"integer","default":5}},"required":["consulta"]}},
    {"name":"enviar_correo","description":"Envía un correo electrónico.","input_schema":{"type":"object","properties":{"destinatario":{"type":"string"},"asunto":{"type":"string"},"cuerpo":{"type":"string"}},"required":["destinatario","asunto","cuerpo"]}},
    {"name":"leer_pdf","description":"Lee y extrae texto de un PDF.","input_schema":{"type":"object","properties":{"ruta_archivo":{"type":"string"}},"required":["ruta_archivo"]}},
    {"name":"leer_excel_csv","description":"Lee archivos Excel o CSV.","input_schema":{"type":"object","properties":{"ruta_archivo":{"type":"string"},"max_filas":{"type":"integer","default":20}},"required":["ruta_archivo"]}},
    {"name":"consultar_base_datos","description":"Ejecuta SQL. Tablas: contactos (id,nombre,email,telefono,notas) y tareas (id,titulo,descripcion,estado,prioridad).","input_schema":{"type":"object","properties":{"consulta_sql":{"type":"string"}},"required":["consulta_sql"]}},
    {"name":"guardar_archivo","description":"Guarda texto en un archivo.","input_schema":{"type":"object","properties":{"nombre_archivo":{"type":"string"},"contenido":{"type":"string"}},"required":["nombre_archivo","contenido"]}},
    {"name":"traducir_texto","description":"Traduce texto entre idiomas.","input_schema":{"type":"object","properties":{"texto":{"type":"string"},"idioma_destino":{"type":"string"},"idioma_origen":{"type":"string","default":"auto"}},"required":["texto","idioma_destino"]}},
    {"name":"generar_grafica","description":"Genera gráficas desde CSV/Excel. Tipos: barras, lineas, pastel, barras_horizontales, dispersión.","input_schema":{"type":"object","properties":{"ruta_csv":{"type":"string"},"tipo_grafica":{"type":"string","default":"barras"},"columna_x":{"type":"string","default":""},"columna_y":{"type":"string","default":""},"titulo":{"type":"string","default":"Gráfica"},"nombre_archivo":{"type":"string","default":"grafica.png"}},"required":["ruta_csv"]}},
    {"name":"resumir_texto","description":"Resume textos largos. Tipos: corto, medio, largo, puntos.","input_schema":{"type":"object","properties":{"texto":{"type":"string"},"longitud":{"type":"string","default":"medio"},"idioma":{"type":"string","default":"español"}},"required":["texto"]}},
    {"name":"obtener_clima","description":"Clima actual de una ciudad.","input_schema":{"type":"object","properties":{"ciudad":{"type":"string"},"pais":{"type":"string","default":"MX"}},"required":["ciudad"]}},
]

FUNCIONES = {
    "buscar_en_web": lambda a: buscar_en_web(a["consulta"], a.get("max_resultados",5)),
    "enviar_correo": lambda a: enviar_correo(a["destinatario"], a["asunto"], a["cuerpo"]),
    "leer_pdf": lambda a: leer_pdf(a["ruta_archivo"]),
    "leer_excel_csv": lambda a: leer_excel_csv(a["ruta_archivo"], a.get("max_filas",20)),
    "consultar_base_datos": lambda a: consultar_base_datos(a["consulta_sql"]),
    "guardar_archivo": lambda a: guardar_archivo(a["nombre_archivo"], a["contenido"]),
    "traducir_texto": lambda a: traducir_texto(a["texto"], a["idioma_destino"], a.get("idioma_origen","auto")),
    "generar_grafica": lambda a: generar_grafica(a["ruta_csv"], a.get("tipo_grafica","barras"), a.get("columna_x",""), a.get("columna_y",""), a.get("titulo","Gráfica"), a.get("nombre_archivo","grafica.png")),
    "resumir_texto": lambda a: resumir_texto(a["texto"], a.get("longitud","medio"), a.get("idioma","español")),
    "obtener_clima": lambda a: obtener_clima(a["ciudad"], a.get("pais","MX")),
}

SYSTEM_PROMPT = """
Eres un asistente personal inteligente que habla español. Tu nombre es "Agente Pro".
Tienes 10 herramientas: buscar_en_web, enviar_correo, leer_pdf, leer_excel_csv,
consultar_base_datos, guardar_archivo, traducir_texto, generar_grafica, resumir_texto, obtener_clima.
Reglas: Responde en español, usa herramientas cuando necesites datos reales, no inventes,
confirma antes de enviar correos, combina herramientas para tareas complejas.
"""


# =============================================================
# INTERFAZ
# =============================================================

# Inicializar DB y estado
inicializar_db()

if "mensajes" not in st.session_state:
    st.session_state.mensajes = []
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []
if "tools_used" not in st.session_state:
    st.session_state.tools_used = 0

# Sidebar
with st.sidebar:
    st.markdown('<div class="sidebar-title">🤖 Agente Pro</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-subtitle">Ultimate Edition — 10 herramientas</div>', unsafe_allow_html=True)

    st.markdown("---")

    tools_info = [
        ("🔍", "Búsqueda web", True),
        ("📧", "Enviar correos", bool(os.getenv("EMAIL_ADDRESS"))),
        ("📄", "Leer PDFs", True),
        ("📊", "Leer Excel/CSV", True),
        ("🗄️", "Base de datos", True),
        ("💾", "Guardar archivos", True),
        ("🌐", "Traductor", True),
        ("📈", "Gráficas", True),
        ("📝", "Resumidor", True),
        ("🌤️", "Clima real", bool(os.getenv("OPENWEATHER_API_KEY"))),
    ]

    for icon, name, active in tools_info:
        status = "● Activo" if active else "○ Config"
        status_class = "status-on" if active else "status-off"
        st.markdown(
            f'<div class="tool-card">'
            f'<span class="tool-icon">{icon}</span>'
            f'<span class="tool-name">{name}</span>'
            f'<span class="tool-status {status_class}">{status}</span>'
            f'</div>',
            unsafe_allow_html=True
        )

    st.markdown("---")

    if st.button("🗑️ Limpiar conversación", use_container_width=True):
        st.session_state.mensajes = []
        st.session_state.chat_history = []
        st.session_state.tools_used = 0
        st.rerun()

    st.markdown("---")
    st.markdown(
        '<div style="font-size: 12px; color: #475569; text-align: center;">'
        'Creado con Claude API + Streamlit<br>by Oswaldo Murillo'
        '</div>',
        unsafe_allow_html=True
    )


# Header principal
st.markdown(
    '<div class="header-container">'
    '<div class="header-icon">🤖</div>'
    '<div class="header-title">Agente Pro Ultimate</div>'
    '<div class="header-desc">Tu asistente personal con 10 herramientas inteligentes</div>'
    '</div>',
    unsafe_allow_html=True
)

# Stats
col1, col2, col3 = st.columns(3)
with col1:
    st.markdown(
        f'<div class="stat-card"><div class="stat-number">{len(st.session_state.chat_history)//2}</div>'
        '<div class="stat-label">Mensajes</div></div>',
        unsafe_allow_html=True
    )
with col2:
    st.markdown(
        f'<div class="stat-card"><div class="stat-number">{st.session_state.tools_used}</div>'
        '<div class="stat-label">Herramientas usadas</div></div>',
        unsafe_allow_html=True
    )
with col3:
    st.markdown(
        '<div class="stat-card"><div class="stat-number">10</div>'
        '<div class="stat-label">Herramientas disponibles</div></div>',
        unsafe_allow_html=True
    )

st.markdown("<br>", unsafe_allow_html=True)

# Mostrar historial de chat
for msg in st.session_state.chat_history:
    if msg["role"] == "user":
        st.markdown(f'<div class="user-msg">{msg["content"]}</div>', unsafe_allow_html=True)
    elif msg["role"] == "assistant":
        st.markdown(f'<div class="agent-msg">{msg["content"]}</div>', unsafe_allow_html=True)
    elif msg["role"] == "tool":
        st.markdown(f'<div class="tool-badge">🔧 {msg["content"]}</div>', unsafe_allow_html=True)

# Mostrar gráficas si se generaron
for msg in st.session_state.chat_history:
    if msg["role"] == "chart":
        if Path(msg["content"]).exists():
            st.image(msg["content"], use_container_width=True)

# Acciones rápidas (solo si no hay mensajes)
if not st.session_state.chat_history:
    st.markdown("<br>", unsafe_allow_html=True)
    cols = st.columns(3)
    quick_actions = [
        "🔍 Busca noticias sobre IA",
        "🗄️ ¿Qué contactos tengo?",
        "🌐 Traduce al inglés: Hola mundo",
        "📝 Resume un texto largo",
        "🌤️ Clima en Guadalajara",
        "📈 Genera gráfica de ventas",
    ]
    for i, action in enumerate(quick_actions):
        with cols[i % 3]:
            if st.button(action, key=f"qa_{i}", use_container_width=True):
                st.session_state.quick_action = action.split(" ", 1)[1]
                st.rerun()

# Input del usuario
pregunta = st.chat_input("Escribe tu mensaje aquí...")

# Verificar si hay quick action pendiente
if "quick_action" in st.session_state:
    pregunta = st.session_state.quick_action
    del st.session_state.quick_action

if pregunta:
    # Agregar mensaje del usuario
    st.session_state.chat_history.append({"role": "user", "content": pregunta})
    st.session_state.mensajes.append({"role": "user", "content": pregunta})

    # Mostrar mensaje del usuario
    st.markdown(f'<div class="user-msg">{pregunta}</div>', unsafe_allow_html=True)

    # Procesar con Claude
    with st.spinner("🤖 Pensando..."):
        client = Anthropic()

        while True:
            try:
                respuesta = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=2048,
                    system=SYSTEM_PROMPT,
                    tools=HERRAMIENTAS,
                    messages=st.session_state.mensajes,
                )
            except Exception as e:
                st.error(f"Error: {str(e)}")
                st.session_state.mensajes.pop()
                break

            if respuesta.stop_reason == "tool_use":
                st.session_state.mensajes.append({"role": "assistant", "content": respuesta.content})
                resultados = []

                for bloque in respuesta.content:
                    if bloque.type == "tool_use":
                        st.session_state.tools_used += 1
                        st.session_state.chat_history.append({"role": "tool", "content": bloque.name})
                        st.markdown(f'<div class="tool-badge">🔧 Usando: {bloque.name}</div>', unsafe_allow_html=True)

                        resultado = FUNCIONES[bloque.name](bloque.input)

                        # Si es gráfica, mostrar imagen
                        if bloque.name == "generar_grafica":
                            archivo = bloque.input.get("nombre_archivo", "grafica.png")
                            if Path(archivo).exists():
                                st.session_state.chat_history.append({"role": "chart", "content": archivo})
                                st.image(archivo, use_container_width=True)

                        resultados.append({
                            "type": "tool_result",
                            "tool_use_id": bloque.id,
                            "content": resultado,
                        })

                st.session_state.mensajes.append({"role": "user", "content": resultados})
            else:
                texto = "".join(b.text for b in respuesta.content if hasattr(b, "text"))
                st.session_state.chat_history.append({"role": "assistant", "content": texto})
                st.session_state.mensajes.append({"role": "assistant", "content": respuesta.content})
                st.markdown(f'<div class="agent-msg">{texto}</div>', unsafe_allow_html=True)
                break

    st.rerun()
