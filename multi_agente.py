"""
=============================================================
🤖 SISTEMA MULTI-AGENTE - Investigador + Escritor
=============================================================

Tres agentes que colaboran:
  1. 🎯 Coordinador - Decide qué agentes usar y orquesta el flujo
  2. 🔍 Investigador - Busca y recopila información
  3. ✍️  Escritor - Redacta documentos pulidos con los datos

Ejemplo de uso:
  "Investiga sobre energía solar en México y escribe un artículo"
  → Coordinador asigna al Investigador → busca datos →
  → pasa resultados al Escritor → genera artículo profesional

INSTALACIÓN:
  pip install anthropic python-dotenv duckduckgo-search PyPDF2 openpyxl

EJECUCIÓN:
  python multi_agente.py

=============================================================
"""

import os
import csv
import time
import json
import sqlite3
from pathlib import Path
from datetime import datetime

from dotenv import load_dotenv
from anthropic import Anthropic

load_dotenv()


# =============================================================
# HERRAMIENTAS COMPARTIDAS
# =============================================================

def buscar_en_web(consulta, max_resultados=5):
    try:
        from duckduckgo_search import DDGS
    except ImportError:
        return "Error: pip install duckduckgo-search"
    for intento in range(3):
        try:
            with DDGS() as ddgs:
                resultados = list(ddgs.text(consulta, max_results=max_resultados, region="wt-wt"))
            if resultados:
                texto = ""
                for i, r in enumerate(resultados, 1):
                    texto += f"{i}. {r['title']}\n   {r['body']}\n   Fuente: {r['href']}\n\n"
                return texto
        except Exception:
            if intento < 2: time.sleep(2)
    return f"Sin resultados para '{consulta}'."


def leer_pdf(ruta_archivo):
    try:
        from PyPDF2 import PdfReader
        ruta = Path(ruta_archivo)
        if not ruta.exists(): return f"No se encontró '{ruta_archivo}'"
        reader = PdfReader(str(ruta))
        texto = ""
        for i, p in enumerate(reader.pages, 1):
            t = p.extract_text()
            if t: texto += f"\n--- Página {i} ---\n{t}"
        if len(texto) > 4000: texto = texto[:4000] + "\n[truncado]"
        return texto if texto.strip() else "PDF sin texto."
    except Exception as e:
        return f"Error: {str(e)}"


def leer_excel_csv(ruta_archivo, max_filas=20):
    try:
        ruta = Path(ruta_archivo)
        if not ruta.exists(): return f"No se encontró '{ruta_archivo}'"
        ext = ruta.suffix.lower()
        if ext == ".csv":
            filas = []
            with open(ruta, "r", encoding="utf-8-sig") as f:
                for i, fila in enumerate(csv.reader(f)):
                    if i >= max_filas + 1: break
                    filas.append(fila)
        elif ext in [".xlsx", ".xls"]:
            from openpyxl import load_workbook
            wb = load_workbook(str(ruta), read_only=True); ws = wb.active
            filas = [[str(c) if c else "" for c in r] for i, r in enumerate(ws.iter_rows(values_only=True)) if i < max_filas + 1]
            wb.close()
        else:
            return f"Formato no soportado: {ext}"
        if not filas: return "Archivo vacío."
        enc, dat = filas[0], filas[1:]
        res = f"Columnas: {', '.join(enc)}\n\n"
        for f in dat: res += " | ".join(f) + "\n"
        return res
    except Exception as e:
        return f"Error: {str(e)}"


def guardar_archivo(nombre, contenido):
    try:
        r = Path(nombre)
        if r.suffix.lower() not in [".txt",".md",".csv",".json",".html"]:
            return "Extensión no permitida."
        with open(r, "w", encoding="utf-8") as f:
            f.write(contenido)
        return f"Archivo '{nombre}' guardado ({len(contenido)} caracteres)."
    except Exception as e:
        return f"Error: {str(e)}"


# =============================================================
# DEFINICIÓN DE AGENTES
# =============================================================

class Agente:
    """Clase base para todos los agentes."""

    def __init__(self, nombre, emoji, system_prompt, herramientas, funciones):
        self.nombre = nombre
        self.emoji = emoji
        self.system_prompt = system_prompt
        self.herramientas = herramientas
        self.funciones = funciones
        self.client = Anthropic()

    def ejecutar(self, tarea, contexto=""):
        """Ejecuta una tarea y devuelve el resultado."""
        print(f"\n  {self.emoji} {self.nombre} trabajando...")

        prompt = tarea
        if contexto:
            prompt = f"CONTEXTO PREVIO (datos de otros agentes):\n{contexto}\n\nTAREA:\n{tarea}"

        mensajes = [{"role": "user", "content": prompt}]

        while True:
            try:
                respuesta = self.client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=2048,
                    system=self.system_prompt,
                    tools=self.herramientas,
                    messages=mensajes,
                )
            except Exception as e:
                return f"Error en {self.nombre}: {str(e)}"

            if respuesta.stop_reason == "tool_use":
                mensajes.append({"role": "assistant", "content": respuesta.content})
                resultados = []

                for bloque in respuesta.content:
                    if bloque.type == "tool_use":
                        nombre_tool = bloque.name
                        print(f"    🔧 {nombre_tool}")

                        if nombre_tool in self.funciones:
                            resultado = self.funciones[nombre_tool](bloque.input)
                        else:
                            resultado = f"Herramienta '{nombre_tool}' no disponible."

                        resultados.append({
                            "type": "tool_result",
                            "tool_use_id": bloque.id,
                            "content": resultado,
                        })

                mensajes.append({"role": "user", "content": resultados})
            else:
                texto = "".join(b.text for b in respuesta.content if hasattr(b, "text"))
                print(f"  ✅ {self.nombre} terminó")
                return texto


# =============================================================
# AGENTE INVESTIGADOR 🔍
# =============================================================

PROMPT_INVESTIGADOR = """
Eres un agente investigador experto. Tu trabajo es buscar, recopilar y
organizar información de forma exhaustiva y precisa.

REGLAS:
1. Siempre busca en múltiples fuentes cuando sea posible
2. Organiza los hallazgos por temas o categorías
3. Incluye datos específicos: números, fechas, nombres, estadísticas
4. Distingue hechos de opiniones
5. Cita las fuentes cuando sea posible
6. Si no encuentras información suficiente, dilo claramente
7. Responde en español

FORMATO DE SALIDA:
Organiza tu investigación así:
- HALLAZGOS PRINCIPALES: Los datos más importantes
- DATOS Y ESTADÍSTICAS: Números relevantes
- FUENTES: De dónde obtuviste la información
- CONTEXTO ADICIONAL: Información de fondo útil
"""

HERRAMIENTAS_INVESTIGADOR = [
    {"name": "buscar_en_web", "description": "Busca información actualizada en internet.",
     "input_schema": {"type": "object", "properties": {
         "consulta": {"type": "string"}, "max_resultados": {"type": "integer", "default": 5}
     }, "required": ["consulta"]}},
    {"name": "leer_pdf", "description": "Lee texto de un PDF.",
     "input_schema": {"type": "object", "properties": {
         "ruta_archivo": {"type": "string"}
     }, "required": ["ruta_archivo"]}},
    {"name": "leer_excel_csv", "description": "Lee archivos Excel o CSV.",
     "input_schema": {"type": "object", "properties": {
         "ruta_archivo": {"type": "string"}, "max_filas": {"type": "integer", "default": 20}
     }, "required": ["ruta_archivo"]}},
]

FUNCIONES_INVESTIGADOR = {
    "buscar_en_web": lambda a: buscar_en_web(a["consulta"], a.get("max_resultados", 5)),
    "leer_pdf": lambda a: leer_pdf(a["ruta_archivo"]),
    "leer_excel_csv": lambda a: leer_excel_csv(a["ruta_archivo"], a.get("max_filas", 20)),
}


# =============================================================
# AGENTE ESCRITOR ✍️
# =============================================================

PROMPT_ESCRITOR = """
Eres un agente escritor profesional. Tu trabajo es tomar datos e información
en bruto y transformarlos en documentos bien escritos, claros y atractivos.

REGLAS:
1. Escribe en español de forma profesional pero accesible
2. Usa una estructura clara con secciones y subtítulos
3. Transforma datos crudos en narrativas coherentes
4. Agrega contexto y conexiones entre los datos
5. Usa un tono apropiado al tipo de documento solicitado
6. Incluye una introducción y conclusión
7. Si te piden guardar, usa la herramienta guardar_archivo

TIPOS DE DOCUMENTOS QUE PUEDES CREAR:
- Artículos informativos
- Reportes ejecutivos
- Resúmenes de investigación
- Publicaciones para blog
- Documentos técnicos
- Emails profesionales
- Presentaciones (en texto)
"""

HERRAMIENTAS_ESCRITOR = [
    {"name": "guardar_archivo", "description": "Guarda el documento final en un archivo.",
     "input_schema": {"type": "object", "properties": {
         "nombre_archivo": {"type": "string"}, "contenido": {"type": "string"}
     }, "required": ["nombre_archivo", "contenido"]}},
    {"name": "traducir_texto", "description": "Traduce texto entre idiomas.",
     "input_schema": {"type": "object", "properties": {
         "texto": {"type": "string"}, "idioma_destino": {"type": "string"}
     }, "required": ["texto", "idioma_destino"]}},
]

def traducir_texto_simple(texto, idioma_destino):
    try:
        client = Anthropic()
        resp = client.messages.create(model="claude-sonnet-4-6", max_tokens=2048,
            messages=[{"role": "user", "content": f"Traduce al {idioma_destino}. Solo la traducción:\n{texto}"}])
        return "".join(b.text for b in resp.content if hasattr(b, "text"))
    except Exception as e:
        return f"Error: {str(e)}"

FUNCIONES_ESCRITOR = {
    "guardar_archivo": lambda a: guardar_archivo(a["nombre_archivo"], a["contenido"]),
    "traducir_texto": lambda a: traducir_texto_simple(a["texto"], a["idioma_destino"]),
}


# =============================================================
# AGENTE COORDINADOR 🎯
# =============================================================

PROMPT_COORDINADOR = """
Eres el agente coordinador de un sistema multi-agente. Tu trabajo es analizar
la solicitud del usuario y crear un plan de trabajo.

AGENTES DISPONIBLES:
1. INVESTIGADOR: Busca información en web, lee PDFs y archivos Excel/CSV.
   Úsalo para: noticias, datos, estadísticas, investigación.
2. ESCRITOR: Redacta documentos profesionales y los guarda en archivos.
   Úsalo para: artículos, reportes, resúmenes, emails.

REGLAS:
1. Analiza la solicitud y decide qué agentes necesitas
2. Responde SOLO con un JSON válido (sin texto adicional, sin backticks)
3. Si solo necesitas investigar, usa solo el investigador
4. Si solo necesitas escribir algo creativo (sin investigación), usa solo el escritor
5. Para tareas complejas, usa ambos: primero investigador, luego escritor

FORMATO DE RESPUESTA (solo JSON, nada más):
{
  "plan": "descripción breve del plan",
  "pasos": [
    {
      "agente": "investigador",
      "tarea": "descripción detallada de lo que debe hacer"
    },
    {
      "agente": "escritor",
      "tarea": "descripción detallada de lo que debe escribir"
    }
  ]
}
"""


# =============================================================
# ORQUESTADOR PRINCIPAL
# =============================================================

class SistemaMultiAgente:
    """Orquesta la colaboración entre agentes."""

    def __init__(self):
        self.client = Anthropic()

        self.investigador = Agente(
            nombre="Investigador",
            emoji="🔍",
            system_prompt=PROMPT_INVESTIGADOR,
            herramientas=HERRAMIENTAS_INVESTIGADOR,
            funciones=FUNCIONES_INVESTIGADOR,
        )

        self.escritor = Agente(
            nombre="Escritor",
            emoji="✍️",
            system_prompt=PROMPT_ESCRITOR,
            herramientas=HERRAMIENTAS_ESCRITOR,
            funciones=FUNCIONES_ESCRITOR,
        )

    def planificar(self, solicitud):
        """El coordinador analiza la solicitud y crea un plan."""
        print("\n  🎯 Coordinador analizando solicitud...")

        try:
            respuesta = self.client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=500,
                system=PROMPT_COORDINADOR,
                messages=[{"role": "user", "content": solicitud}],
            )

            texto = "".join(b.text for b in respuesta.content if hasattr(b, "text"))

            # Limpiar posibles backticks
            texto = texto.strip()
            if texto.startswith("```"):
                texto = texto.split("\n", 1)[1] if "\n" in texto else texto[3:]
            if texto.endswith("```"):
                texto = texto[:-3]
            texto = texto.strip()

            plan = json.loads(texto)
            print(f"  📋 Plan: {plan['plan']}")
            print(f"  📊 Pasos: {len(plan['pasos'])}")
            return plan

        except json.JSONDecodeError as e:
            print(f"  ⚠️ Error parseando plan, usando modo directo")
            # Fallback: si no puede parsear, envía todo al investigador + escritor
            return {
                "plan": "Investigar y escribir sobre el tema",
                "pasos": [
                    {"agente": "investigador", "tarea": f"Investiga sobre: {solicitud}"},
                    {"agente": "escritor", "tarea": f"Con la investigación anterior, escribe un documento sobre: {solicitud}"}
                ]
            }
        except Exception as e:
            print(f"  ❌ Error: {str(e)}")
            return None

    def ejecutar(self, solicitud):
        """Ejecuta el flujo completo: planificar → investigar → escribir."""

        # Paso 1: Planificar
        plan = self.planificar(solicitud)
        if not plan:
            return "Error al crear el plan de trabajo."

        # Paso 2: Ejecutar cada paso del plan
        contexto_acumulado = ""
        resultado_final = ""

        for i, paso in enumerate(plan["pasos"], 1):
            agente_nombre = paso["agente"].lower()
            tarea = paso["tarea"]

            print(f"\n  {'─' * 40}")
            print(f"  Paso {i}/{len(plan['pasos'])}: {agente_nombre.upper()}")
            print(f"  Tarea: {tarea[:80]}...")

            if agente_nombre == "investigador":
                resultado = self.investigador.ejecutar(tarea, contexto_acumulado)
                contexto_acumulado += f"\n\nRESULTADOS DEL INVESTIGADOR:\n{resultado}"

            elif agente_nombre == "escritor":
                resultado = self.escritor.ejecutar(tarea, contexto_acumulado)
                resultado_final = resultado

            else:
                print(f"  ⚠️ Agente '{agente_nombre}' no reconocido, saltando...")
                continue

        # Si no hubo escritor, el resultado es lo que recopiló el investigador
        if not resultado_final:
            resultado_final = contexto_acumulado

        return resultado_final


# =============================================================
# INTERFAZ DE TERMINAL
# =============================================================

def main():
    sistema = SistemaMultiAgente()

    print("")
    print("=" * 58)
    print("  🤖 SISTEMA MULTI-AGENTE")
    print("  Investigador + Escritor colaborando")
    print("=" * 58)
    print("")
    print("  Agentes disponibles:")
    print("  🔍 Investigador — busca y analiza datos")
    print("  ✍️  Escritor — redacta documentos profesionales")
    print("  🎯 Coordinador — orquesta la colaboración")
    print("")
    print("  Ejemplos de tareas:")
    print('  • "Investiga sobre IA en México y escribe un artículo"')
    print('  • "Busca las últimas noticias de tecnología y')
    print('     hazme un resumen ejecutivo"')
    print('  • "Investiga sobre energía solar y escribe un')
    print('     reporte guardándolo en energia_solar.md"')
    print('  • "Escribe un email profesional presentando mi')
    print('     empresa de software"')
    print("")
    print('  Escribe "salir" para terminar.')
    print("=" * 58)

    while True:
        solicitud = input("\n🧑 Tú: ").strip()
        if solicitud.lower() in ["salir", "exit", "quit"]:
            print("\n👋 ¡Hasta luego!")
            break
        if not solicitud:
            continue

        print("\n" + "=" * 50)
        print("  🚀 Iniciando sistema multi-agente...")
        print("=" * 50)

        resultado = sistema.ejecutar(solicitud)

        print(f"\n{'=' * 50}")
        print("  📄 RESULTADO FINAL")
        print(f"{'=' * 50}")
        print(f"\n{resultado}")


if __name__ == "__main__":
    main()
