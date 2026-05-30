"""
=============================================================
🤖 AGENTE PRO - Bot de Telegram con Memoria
=============================================================

Usa tu agente con 13 herramientas + memoria a largo plazo.
El agente recuerda conversaciones anteriores incluso
después de reiniciarlo.

INSTALACIÓN:
  pip install python-telegram-bot anthropic python-dotenv duckduckgo-search PyPDF2 openpyxl matplotlib requests

ARCHIVOS NECESARIOS:
  - telegram_bot_memoria.py  (este archivo)
  - memoria.py               (módulo de memoria)
  - .env                     (configuración)

EJECUCIÓN:
  python telegram_bot_memoria.py

=============================================================
"""

import os
import json
import sqlite3
import smtplib
import csv
import time
import logging
import threading
from http.server import HTTPServer, BaseHTTPRequestHandler
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from anthropic import Anthropic
from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
)

load_dotenv()

# Importar sistema de memoria
from memoria import (
    HERRAMIENTAS_MEMORIA, FUNCIONES_MEMORIA,
    obtener_system_prompt_con_memoria,
    agregar_intercambio, verificar_actualizacion_resumen,
    listar_hechos
)
from alertas import iniciar_sistema_alertas

# Configurar logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)


# =============================================================
# TODAS LAS HERRAMIENTAS (mismas del agente_pro_ultimate.py)
# =============================================================

def buscar_en_web(consulta, max_resultados=5):
    try:
        from duckduckgo_search import DDGS
    except ImportError:
        return "Error: duckduckgo-search no instalado"
    for intento in range(3):
        try:
            with DDGS() as ddgs:
                resultados = list(ddgs.text(consulta, max_results=max_resultados, region="wt-wt"))
            if resultados:
                texto = f"Resultados para '{consulta}':\n\n"
                for i, r in enumerate(resultados, 1):
                    texto += f"{i}. {r['title']}\n   {r['body']}\n   {r['href']}\n\n"
                return texto
        except Exception:
            if intento < 2: time.sleep(2)
    try:
        with DDGS() as ddgs:
            resultados = list(ddgs.news(consulta, max_results=max_resultados))
        if resultados:
            texto = f"Noticias para '{consulta}':\n\n"
            for i, r in enumerate(resultados, 1):
                texto += f"{i}. {r['title']}\n   {r['body']}\n   {r['source']}\n\n"
            return texto
    except Exception:
        pass
    return f"No se encontraron resultados para '{consulta}'."


def enviar_correo(destinatario, asunto, cuerpo):
    ea, ep = os.getenv("EMAIL_ADDRESS"), os.getenv("EMAIL_PASSWORD")
    if not ea or not ep:
        return "Error: Configura EMAIL_ADDRESS y EMAIL_PASSWORD en .env"
    try:
        msg = MIMEMultipart()
        msg["From"], msg["To"], msg["Subject"] = ea, destinatario, asunto
        msg.attach(MIMEText(cuerpo, "plain"))
        with smtplib.SMTP("smtp.gmail.com", 587) as s:
            s.starttls(); s.login(ea, ep); s.send_message(msg)
        return f"Correo enviado a {destinatario}"
    except Exception as e:
        return f"Error: {str(e)}"


def leer_pdf(ruta_archivo):
    try:
        from PyPDF2 import PdfReader
    except ImportError:
        return "Error: PyPDF2 no instalado"
    try:
        ruta = Path(ruta_archivo)
        if not ruta.exists(): return f"No se encontró '{ruta_archivo}'"
        reader = PdfReader(str(ruta))
        texto = ""
        for i, p in enumerate(reader.pages, 1):
            t = p.extract_text()
            if t: texto += f"\n--- Página {i} ---\n{t}"
        if not texto.strip(): return "PDF sin texto extraíble."
        if len(texto) > 3000: texto = texto[:3000] + "\n[... truncado ...]"
        return f"PDF '{ruta.name}' ({len(reader.pages)} págs):\n{texto}"
    except Exception as e:
        return f"Error: {str(e)}"


def leer_excel_csv(ruta_archivo, max_filas=20):
    try:
        ruta = Path(ruta_archivo)
        if not ruta.exists(): return f"No se encontró '{ruta_archivo}'"
        ext = ruta.suffix.lower()
        if ext == ".csv":
            filas = []
            with open(ruta, "r", encoding="utf-8-sig") as f:
                for i, fila in enumerate(csv.reader(f)):
                    if i >= max_filas + 1: break
                    filas.append(fila)
        elif ext in [".xlsx", ".xls"]:
            from openpyxl import load_workbook
            wb = load_workbook(str(ruta), read_only=True); ws = wb.active
            filas = [[str(c) if c else "" for c in r] for i, r in enumerate(ws.iter_rows(values_only=True)) if i < max_filas + 1]
            wb.close()
        else:
            return f"Formato no soportado: {ext}"
        if not filas: return "Archivo vacío."
        enc, dat = filas[0], filas[1:]
        res = f"Archivo: {ruta.name}\nColumnas: {', '.join(enc)}\n\n"
        for f in dat: res += " | ".join(f) + "\n"
        return res
    except Exception as e:
        return f"Error: {str(e)}"


import platform
if platform.system() == "Windows":
    DB_PATH = r"C:\Users\omurillo\.gemini\antigravity\scratch\claude_agent\agente_datos.db"
else:
    DB_PATH = "agente_datos.db"

def inicializar_db():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS contactos (id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL, email TEXT, telefono TEXT, notas TEXT, creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    c.execute("CREATE TABLE IF NOT EXISTS tareas (id INTEGER PRIMARY KEY AUTOINCREMENT, titulo TEXT NOT NULL, descripcion TEXT, estado TEXT DEFAULT 'pendiente', prioridad TEXT DEFAULT 'media', creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
    c.execute("SELECT COUNT(*) FROM contactos")
    if c.fetchone()[0] == 0:
        c.executemany("INSERT INTO contactos (nombre, email, telefono, notas) VALUES (?, ?, ?, ?)",
            [("Ana García","ana@ejemplo.com","33-1234-5678","Cliente principal"),
             ("Carlos López","carlos@ejemplo.com","81-9876-5432","Proveedor"),
             ("María Rodríguez","maria@ejemplo.com","55-5555-1234","Socia")])
    c.execute("SELECT COUNT(*) FROM tareas")
    if c.fetchone()[0] == 0:
        c.executemany("INSERT INTO tareas (titulo, descripcion, estado, prioridad) VALUES (?, ?, ?, ?)",
            [("Revisar propuesta","Propuesta del cliente","pendiente","alta"),
             ("Enviar reporte","Reporte mensual","pendiente","media"),
             ("Llamar a proveedor","Negociar precios","completada","baja")])
    conn.commit(); conn.close()


def consultar_base_datos(sql):
    """Usa Supabase si está configurado, si no usa SQLite local."""
    if os.environ.get("SUPABASE_URL"):
        from db_supabase import consultar_sql
        return consultar_sql(sql)
    
    up = sql.strip().upper()
    if any(p in up for p in ["DROP","DELETE","ALTER","TRUNCATE"]): return "Operación no permitida."
    try:
        conn = sqlite3.connect(DB_PATH); c = conn.cursor(); c.execute(sql)
        if up.startswith("SELECT"):
            cols = [d[0] for d in c.description] if c.description else []
            filas = c.fetchall()
            if not filas: res = "Sin resultados."
            else:
                res = f"Columnas: {', '.join(cols)}\n\n"
                for f in filas: res += " | ".join(str(v) for v in f) + "\n"
        else:
            conn.commit(); res = f"Ejecutado. Filas afectadas: {c.rowcount}"
        conn.close(); return res
    except Exception as e:
        return f"Error: {str(e)}"


def guardar_archivo(nombre, contenido):
    try:
        r = Path(nombre)
        if r.suffix.lower() not in [".txt",".md",".csv",".json",".html"]: return "Extensión no permitida."
        with open(r, "w", encoding="utf-8") as f: f.write(contenido)
        return f"Archivo '{nombre}' guardado ({len(contenido)} chars)."
    except Exception as e:
        return f"Error: {str(e)}"


def traducir_texto(texto, idioma_destino, idioma_origen="auto"):
    try:
        client = Anthropic()
        resp = client.messages.create(model="claude-sonnet-4-6", max_tokens=2048,
            messages=[{"role":"user","content":f"Traduce al {idioma_destino}. Solo la traducción:\n{texto}"}])
        trad = "".join(b.text for b in resp.content if hasattr(b,"text"))
        return f"Traducción al {idioma_destino}:\n\n{trad}"
    except Exception as e:
        return f"Error: {str(e)}"


def generar_grafica(ruta_csv, tipo="barras", col_x="", col_y="", titulo="Gráfica", archivo="grafica.png"):
    try:
        import matplotlib; matplotlib.use("Agg")
        import matplotlib.pyplot as plt; import matplotlib.ticker as ticker
    except ImportError:
        return "Error: matplotlib no instalado"
    try:
        ruta = Path(ruta_csv)
        if not ruta.exists(): return f"No se encontró '{ruta_csv}'"
        datos, enc = [], []
        if ruta.suffix.lower() == ".csv":
            with open(ruta,"r",encoding="utf-8-sig") as f:
                r = csv.reader(f); enc = next(r); datos = list(r)
        elif ruta.suffix.lower() in [".xlsx",".xls"]:
            from openpyxl import load_workbook
            wb = load_workbook(str(ruta),read_only=True); ws = wb.active
            filas = list(ws.iter_rows(values_only=True)); wb.close()
            enc = [str(c) if c else f"C{i}" for i,c in enumerate(filas[0])]
            datos = [[str(c) if c else "" for c in f] for f in filas[1:]]
        if not col_x: col_x = enc[0]
        if not col_y:
            for i,h in enumerate(enc[1:],1):
                try: float(datos[0][i].replace(",","").replace("$","")); col_y=h; break
                except: continue
            if not col_y: col_y = enc[1] if len(enc)>1 else enc[0]
        xi = yi = None
        for i,h in enumerate(enc):
            if h.strip().lower()==col_x.strip().lower(): xi=i
            if h.strip().lower()==col_y.strip().lower(): yi=i
        if xi is None: return f"Columna '{col_x}' no encontrada."
        if yi is None: return f"Columna '{col_y}' no encontrada."
        ag = {}
        for f in datos:
            try:
                xv = f[xi].strip()
                yv = float(f[yi].strip().replace(",","").replace("$",""))
                ag[xv] = ag.get(xv,0)+yv
            except: continue
        if not ag: return "Sin datos numéricos."
        ex, vy = list(ag.keys()), list(ag.values())
        plt.style.use("seaborn-v0_8-whitegrid"); fig, ax = plt.subplots(figsize=(10,6))
        cols = ["#378ADD","#1D9E75","#D85A30","#D4537E","#BA7517","#534AB7"]
        t = tipo.lower()
        if t in ["barras","bar"]: ax.bar(ex,vy,color=cols[:len(ex)],edgecolor="white")
        elif t in ["lineas","linea","line"]: ax.plot(ex,vy,color=cols[0],marker="o",linewidth=2)
        elif t in ["pastel","pie"]: ax.pie(vy,labels=ex,colors=cols[:len(ex)],autopct='%1.1f%%'); ax.axis('equal')
        elif t in ["barras_horizontales","barh"]: ax.barh(ex,vy,color=cols[:len(ex)])
        else: ax.scatter(range(len(vy)),vy,c=cols[0],s=100)
        ax.set_title(titulo,fontsize=16,fontweight="bold")
        plt.tight_layout()
        if Path(archivo).suffix.lower() not in [".png",".jpg"]: archivo+=".png"
        plt.savefig(archivo,dpi=150,bbox_inches='tight',facecolor='white'); plt.close()
        return f"GRAFICA_GENERADA:{archivo}"
    except Exception as e:
        return f"Error: {str(e)}"


def resumir_texto(texto, longitud="medio", idioma="español"):
    try:
        inst = {"corto":"3 oraciones","medio":"1-2 párrafos","largo":"detallado","puntos":"5-7 viñetas"}
        client = Anthropic()
        resp = client.messages.create(model="claude-sonnet-4-6", max_tokens=2048,
            messages=[{"role":"user","content":f"Resume en {idioma}. Formato: {inst.get(longitud,'1-2 párrafos')}.\n\n{texto}"}])
        return "".join(b.text for b in resp.content if hasattr(b,"text"))
    except Exception as e:
        return f"Error: {str(e)}"


def obtener_clima(ciudad, pais="MX"):
    ak = os.getenv("OPENWEATHER_API_KEY")
    if not ak: return "Configura OPENWEATHER_API_KEY en .env (gratis en openweathermap.org)"
    try:
        import requests
        r = requests.get("https://api.openweathermap.org/data/2.5/weather",
            params={"q":f"{ciudad},{pais}","appid":ak,"units":"metric","lang":"es"},timeout=10)
        d = r.json()
        if r.status_code!=200: return f"Error: {d.get('message','Ciudad no encontrada')}"
        return (f"🌤️ Clima en {d['name']}, {d['sys']['country']}:\n\n"
                f"🌡️ {d['main']['temp']:.1f}°C (sensación {d['main']['feels_like']:.1f}°C)\n"
                f"📊 Mín: {d['main']['temp_min']:.1f}°C | Máx: {d['main']['temp_max']:.1f}°C\n"
                f"☁️ {d['weather'][0]['description'].capitalize()}\n"
                f"💧 Humedad: {d['main']['humidity']}%\n"
                f"💨 Viento: {d['wind']['speed']} m/s")
    except Exception as e:
        return f"Error: {str(e)}"


# =============================================================
# CONFIGURACIÓN DE HERRAMIENTAS PARA CLAUDE
# =============================================================

HERRAMIENTAS = [
    {"name":"buscar_en_web","description":"Busca información actualizada en internet.","input_schema":{"type":"object","properties":{"consulta":{"type":"string"},"max_resultados":{"type":"integer","default":5}},"required":["consulta"]}},
    {"name":"enviar_correo","description":"Envía un correo electrónico.","input_schema":{"type":"object","properties":{"destinatario":{"type":"string"},"asunto":{"type":"string"},"cuerpo":{"type":"string"}},"required":["destinatario","asunto","cuerpo"]}},
    {"name":"leer_pdf","description":"Lee texto de un PDF.","input_schema":{"type":"object","properties":{"ruta_archivo":{"type":"string"}},"required":["ruta_archivo"]}},
    {"name":"leer_excel_csv","description":"Lee archivos Excel o CSV.","input_schema":{"type":"object","properties":{"ruta_archivo":{"type":"string"},"max_filas":{"type":"integer","default":20}},"required":["ruta_archivo"]}},
    {"name":"consultar_base_datos","description":"SQL en tablas: contactos (id,nombre,email,telefono,notas) y tareas (id,titulo,descripcion,estado,prioridad).","input_schema":{"type":"object","properties":{"consulta_sql":{"type":"string"}},"required":["consulta_sql"]}},
    {"name":"guardar_archivo","description":"Guarda texto en archivo.","input_schema":{"type":"object","properties":{"nombre_archivo":{"type":"string"},"contenido":{"type":"string"}},"required":["nombre_archivo","contenido"]}},
    {"name":"traducir_texto","description":"Traduce texto entre idiomas.","input_schema":{"type":"object","properties":{"texto":{"type":"string"},"idioma_destino":{"type":"string"},"idioma_origen":{"type":"string","default":"auto"}},"required":["texto","idioma_destino"]}},
    {"name":"generar_grafica","description":"Gráficas desde CSV/Excel: barras, lineas, pastel.","input_schema":{"type":"object","properties":{"ruta_csv":{"type":"string"},"tipo_grafica":{"type":"string","default":"barras"},"columna_x":{"type":"string","default":""},"columna_y":{"type":"string","default":""},"titulo":{"type":"string","default":"Gráfica"},"nombre_archivo":{"type":"string","default":"grafica.png"}},"required":["ruta_csv"]}},
    {"name":"resumir_texto","description":"Resume textos largos (corto, medio, largo, puntos).","input_schema":{"type":"object","properties":{"texto":{"type":"string"},"longitud":{"type":"string","default":"medio"},"idioma":{"type":"string","default":"español"}},"required":["texto"]}},
    {"name":"obtener_clima","description":"Clima actual de una ciudad.","input_schema":{"type":"object","properties":{"ciudad":{"type":"string"},"pais":{"type":"string","default":"MX"}},"required":["ciudad"]}},
] + HERRAMIENTAS_MEMORIA  # Agregar herramientas de memoria

FUNCIONES = {
    "buscar_en_web": lambda a: buscar_en_web(a["consulta"], a.get("max_resultados",5)),
    "enviar_correo": lambda a: enviar_correo(a["destinatario"], a["asunto"], a["cuerpo"]),
    "leer_pdf": lambda a: leer_pdf(a["ruta_archivo"]),
    "leer_excel_csv": lambda a: leer_excel_csv(a["ruta_archivo"], a.get("max_filas",20)),
    "consultar_base_datos": lambda a: consultar_base_datos(a["consulta_sql"]),
    "guardar_archivo": lambda a: guardar_archivo(a["nombre_archivo"], a["contenido"]),
    "traducir_texto": lambda a: traducir_texto(a["texto"], a["idioma_destino"], a.get("idioma_origen","auto")),
    "generar_grafica": lambda a: generar_grafica(a["ruta_csv"], a.get("tipo_grafica","barras"), a.get("columna_x",""), a.get("columna_y",""), a.get("titulo","Gráfica"), a.get("nombre_archivo","grafica.png")),
    "resumir_texto": lambda a: resumir_texto(a["texto"], a.get("longitud","medio"), a.get("idioma","español")),
    "obtener_clima": lambda a: obtener_clima(a["ciudad"], a.get("pais","MX")),
    **FUNCIONES_MEMORIA,  # Agregar funciones de memoria
}

SYSTEM_PROMPT = """
Eres un asistente personal inteligente en Telegram. Tu nombre es "Agente Pro".
Hablas español de forma amigable. Tienes 13 herramientas disponibles.

HERRAMIENTAS PRINCIPALES:
buscar_en_web, enviar_correo, leer_pdf, leer_excel_csv, consultar_base_datos,
guardar_archivo, traducir_texto, generar_grafica, resumir_texto, obtener_clima.

HERRAMIENTAS DE MEMORIA:
- recordar_hecho: Guarda datos importantes del usuario (nombre, preferencias, etc.)
- olvidar_hecho: Borra algo de la memoria
- listar_memoria: Muestra todo lo que recuerdas

REGLAS:
1. Responde en español, de forma concisa
2. Usa emojis para hacer las respuestas más visuales
3. Usa herramientas cuando necesites datos reales
4. Confirma antes de enviar correos
5. Combina herramientas para tareas complejas
6. Cuando el usuario comparta info personal importante (nombre, trabajo, gustos),
   usa recordar_hecho para guardarla automáticamente
7. Usa la memoria para personalizar tus respuestas
8. Nunca digas "según mi memoria" o "en mis registros" - simplemente usa la info naturalmente
"""


# =============================================================
# HISTORIAL DE CONVERSACIONES POR USUARIO
# =============================================================

# Cada usuario de Telegram tiene su propio historial
conversaciones = {}
MAX_HISTORIAL = 20  # Máximo de mensajes por usuario


def obtener_historial(user_id):
    """Obtiene o crea el historial de un usuario."""
    if user_id not in conversaciones:
        conversaciones[user_id] = []
    return conversaciones[user_id]


def limpiar_historial(user_id):
    """Limpia el historial de un usuario."""
    conversaciones[user_id] = []


# =============================================================
# PROCESADOR DE MENSAJES CON CLAUDE
# =============================================================

def procesar_con_claude(user_id, mensaje_texto):
    """Envía un mensaje a Claude y procesa la respuesta con herramientas + memoria."""
    client = Anthropic()
    historial = obtener_historial(user_id)

    # Agregar mensaje del usuario
    historial.append({"role": "user", "content": mensaje_texto})

    # Limitar historial para no exceder contexto
    if len(historial) > MAX_HISTORIAL:
        historial = historial[-MAX_HISTORIAL:]
        conversaciones[user_id] = historial

    # Enriquecer system prompt con memoria
    system_con_memoria = obtener_system_prompt_con_memoria(SYSTEM_PROMPT)

    herramientas_usadas = []
    grafica_generada = None

    # Loop del agente
    while True:
        try:
            respuesta = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=1500,
                system=system_con_memoria,
                tools=HERRAMIENTAS,
                messages=historial,
            )
        except Exception as e:
            historial.pop()
            return f"❌ Error al conectar con Claude: {str(e)}", None

        if respuesta.stop_reason == "tool_use":
            historial.append({"role": "assistant", "content": respuesta.content})
            resultados = []

            for bloque in respuesta.content:
                if bloque.type == "tool_use":
                    nombre = bloque.name
                    herramientas_usadas.append(nombre)
                    logger.info(f"🔧 Usuario {user_id} usando: {nombre}")

                    resultado = FUNCIONES[nombre](bloque.input)

                    # Detectar si se generó una gráfica
                    if resultado.startswith("GRAFICA_GENERADA:"):
                        grafica_generada = resultado.replace("GRAFICA_GENERADA:", "")
                        resultado = f"Gráfica generada exitosamente en {grafica_generada}"

                    resultados.append({
                        "type": "tool_result",
                        "tool_use_id": bloque.id,
                        "content": resultado,
                    })

            historial.append({"role": "user", "content": resultados})
        else:
            # Respuesta final
            texto = "".join(b.text for b in respuesta.content if hasattr(b, "text"))
            historial.append({"role": "assistant", "content": respuesta.content})

            # Guardar intercambio en memoria persistente
            agregar_intercambio(mensaje_texto, texto)
            verificar_actualizacion_resumen()

            # Agregar indicador de herramientas usadas
            if herramientas_usadas:
                tools_text = " | ".join(f"🔧{t}" for t in herramientas_usadas)
                texto = f"{texto}\n\n───────────\n{tools_text}"

            return texto, grafica_generada


# =============================================================
# HANDLERS DE TELEGRAM
# =============================================================

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando /start - Bienvenida."""
    user = update.effective_user
    bienvenida = (
        f"👋 ¡Hola {user.first_name}! Soy tu **Agente Pro**.\n\n"
        f"Tengo 13 herramientas para ayudarte:\n\n"
        f"🔍 Búsqueda web    📧 Correos\n"
        f"📄 Leer PDFs       📊 Excel/CSV\n"
        f"🗄️ Base de datos   💾 Archivos\n"
        f"🌐 Traductor       📈 Gráficas\n"
        f"📝 Resumidor       🌤️ Clima\n"
        f"🧠 Memoria a largo plazo\n\n"
        f"Ahora tengo **memoria**: recuerdo lo que me dices\n"
        f"incluso después de reiniciarme.\n\n"
        f"Comandos:\n"
        f"/start - Ver este mensaje\n"
        f"/memoria - Ver qué recuerdo de ti\n"
        f"/limpiar - Borrar historial de sesión\n"
        f"/herramientas - Ver herramientas\n"
        f"/estado - Ver estado del agente"
        f"/alertas - Activar recordatorios\n"
    )
    # Iniciar sistema de alertas para este usuario
    chat_id = update.effective_chat.id
    bot_token = os.getenv("TELEGRAM_BOT_TOKEN")
    if bot_token and not hasattr(cmd_start, '_alertas_iniciadas'):
        iniciar_sistema_alertas(bot_token, chat_id)
        cmd_start._alertas_iniciadas = True

    await update.message.reply_text(bienvenida, parse_mode="Markdown")


async def cmd_limpiar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando /limpiar - Borrar historial."""
    user_id = update.effective_user.id
    limpiar_historial(user_id)
    await update.message.reply_text("🧹 Historial limpiado. ¡Empecemos de nuevo!")


async def cmd_herramientas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando /herramientas - Lista de herramientas."""
    texto = (
        "🛠️ **Herramientas disponibles:**\n\n"
        "🔍 **Búsqueda web** - Info actualizada de internet\n"
        "📧 **Correo** - Enviar emails desde Gmail\n"
        "📄 **Leer PDF** - Extraer texto de PDFs\n"
        "📊 **Excel/CSV** - Leer hojas de cálculo\n"
        "🗄️ **Base de datos** - Contactos y tareas\n"
        "💾 **Archivos** - Guardar resultados\n"
        "🌐 **Traductor** - Cualquier idioma\n"
        "📈 **Gráficas** - Barras, líneas, pastel\n"
        "📝 **Resumidor** - Resumir textos largos\n"
        "🌤️ **Clima** - Temperatura en tiempo real"
    )
    await update.message.reply_text(texto, parse_mode="Markdown")


async def cmd_estado(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando /estado - Estado del agente."""
    user_id = update.effective_user.id
    historial = obtener_historial(user_id)
    n_mensajes = len([m for m in historial if m["role"] == "user"])

    email_ok = "✅" if os.getenv("EMAIL_ADDRESS") else "❌"
    clima_ok = "✅" if os.getenv("OPENWEATHER_API_KEY") else "❌"

    texto = (
        f"📊 **Estado del Agente Pro**\n\n"
        f"🤖 Modelo: Claude Sonnet 4\n"
        f"💬 Mensajes en sesión: {n_mensajes}\n"
        f"🔧 Herramientas: 13\n"
        f"🧠 Memoria: Activa\n\n"
        f"**Configuración:**\n"
        f"🔑 API Anthropic: ✅\n"
        f"📧 Email (Gmail): {email_ok}\n"
        f"🌤️ Clima (OpenWeather): {clima_ok}\n"
    )
    await update.message.reply_text(texto, parse_mode="Markdown")


async def cmd_memoria(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando /memoria - Ver qué recuerda el agente."""
    hechos = listar_hechos()
    await update.message.reply_text(f"🧠 **Mi memoria:**\n\n{hechos}", parse_mode="Markdown")


async def manejar_mensaje(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Maneja mensajes de texto del usuario."""
    user_id = update.effective_user.id
    mensaje = update.message.text

    logger.info(f"📩 {update.effective_user.first_name} ({user_id}): {mensaje[:50]}...")

    # Enviar indicador de "escribiendo..."
    await update.message.chat.send_action("typing")

    # Procesar con Claude
    respuesta, grafica = procesar_con_claude(user_id, mensaje)

    # Telegram tiene límite de 4096 caracteres por mensaje
    if len(respuesta) > 4000:
        # Dividir en partes
        partes = [respuesta[i:i+4000] for i in range(0, len(respuesta), 4000)]
        for parte in partes:
            await update.message.reply_text(parte)
    else:
        await update.message.reply_text(respuesta)

    # Si se generó una gráfica, enviarla como imagen
    if grafica and Path(grafica).exists():
        try:
            with open(grafica, "rb") as img:
                await update.message.reply_photo(photo=img, caption="📈 Gráfica generada")
        except Exception as e:
            logger.error(f"Error enviando gráfica: {e}")


async def manejar_documento(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Maneja documentos enviados por el usuario (PDFs, CSV, Excel)."""
    doc = update.message.document
    file_name = doc.file_name
    ext = Path(file_name).suffix.lower()

    if ext not in [".pdf", ".csv", ".xlsx", ".xls", ".txt"]:
        await update.message.reply_text(
            f"⚠️ Formato {ext} no soportado. Envía archivos PDF, CSV, Excel o TXT."
        )
        return

    # Descargar archivo
    await update.message.chat.send_action("typing")
    file = await context.bot.get_file(doc.file_id)
    ruta_local = Path(f"telegram_files/{file_name}")
    ruta_local.parent.mkdir(exist_ok=True)
    await file.download_to_drive(str(ruta_local))

    # Procesar con el agente
    mensaje = f"He subido el archivo {file_name}. Está en la ruta: {ruta_local.absolute()}"

    if update.message.caption:
        mensaje += f"\n\nEl usuario dice: {update.message.caption}"
    else:
        if ext == ".pdf":
            mensaje += "\nPor favor lee y resume este PDF."
        elif ext in [".csv", ".xlsx", ".xls"]:
            mensaje += "\nPor favor lee y analiza este archivo de datos."
        else:
            mensaje += "\nPor favor lee este archivo."

    respuesta, grafica = procesar_con_claude(update.effective_user.id, mensaje)

    if len(respuesta) > 4000:
        partes = [respuesta[i:i+4000] for i in range(0, len(respuesta), 4000)]
        for parte in partes:
            await update.message.reply_text(parte)
    else:
        await update.message.reply_text(respuesta)

    if grafica and Path(grafica).exists():
        with open(grafica, "rb") as img:
            await update.message.reply_photo(photo=img, caption="📈 Gráfica generada")


# =============================================================
# PUNTO DE ENTRADA
# =============================================================

def main():
    """Inicia el bot de Telegram."""
    token = os.getenv("TELEGRAM_BOT_TOKEN")

    if not token:
        print("❌ Error: Configura TELEGRAM_BOT_TOKEN en tu archivo .env")
        print("   Obtén uno hablándole a @BotFather en Telegram.")
        return

    if not os.getenv("ANTHROPIC_API_KEY"):
        print("❌ Error: Configura ANTHROPIC_API_KEY en tu archivo .env")
        return

    # Inicializar base de datos
    inicializar_db()

    # Crear carpeta para archivos de Telegram
    Path("telegram_files").mkdir(exist_ok=True)

    print("")
    print("=" * 50)
    print("  🤖 AGENTE PRO - Bot de Telegram con Memoria")
    print("=" * 50)
    print("")
    print("  ✅ Bot iniciado correctamente")
    print("  📱 Abre Telegram y busca tu bot")
    print("  💬 Envíale un mensaje para empezar")
    print("")
    print("  Presiona Ctrl+C para detener el bot")
    print("=" * 50)

    # Crear aplicación
    app = Application.builder().token(token).build()

    # Registrar handlers
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("limpiar", cmd_limpiar))
    app.add_handler(CommandHandler("herramientas", cmd_herramientas))
    app.add_handler(CommandHandler("estado", cmd_estado))
    app.add_handler(CommandHandler("memoria", cmd_memoria))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, manejar_mensaje))
    app.add_handler(MessageHandler(filters.Document.ALL, manejar_documento))
# Servidor web para Render health check
    class HealthHandler(BaseHTTPRequestHandler):
        def do_GET(self):
            self.send_response(200)
            self.end_headers()
            self.wfile.write(b"Bot activo")
        def log_message(self, format, *args):
            pass
    
    port = int(os.environ.get("PORT", 10000))
    server = HTTPServer(("0.0.0.0", port), HealthHandler)
    threading.Thread(target=server.serve_forever, daemon=True).start()
    print(f"  🌐 Health check en puerto {port}")
    # Iniciar bot
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
