"""
=============================================================
🤖 AGENTE PRO ULTIMATE - Todas las Herramientas
=============================================================

10 herramientas reales:
  1.  🔍 Buscar en la web (DuckDuckGo)
  2.  📧 Enviar correos (Gmail SMTP)
  3.  📄 Leer PDFs
  4.  📊 Leer Excel/CSV
  5.  🗄️  Base de datos SQLite
  6.  💾 Guardar archivos
  7.  🌐 Traductor de idiomas
  8.  📈 Generar gráficas
  9.  📝 Resumidor de textos
  10. 🌤️  Clima en tiempo real (OpenWeatherMap)

INSTALACIÓN:
  pip install anthropic python-dotenv duckduckgo-search PyPDF2 openpyxl matplotlib requests

CONFIGURACIÓN (.env):
  ANTHROPIC_API_KEY=sk-ant-tu-api-key
  EMAIL_ADDRESS=tu-correo@gmail.com          (opcional)
  EMAIL_PASSWORD=xxxx-xxxx-xxxx-xxxx         (opcional - contraseña de app Google)
  OPENWEATHER_API_KEY=tu-key-de-openweather  (opcional - gratis en openweathermap.org)

=============================================================
"""

import os
import json
import sqlite3
import smtplib
import csv
import time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from anthropic import Anthropic

load_dotenv()


# =============================================================
# 1. 🔍 BÚSQUEDA WEB (DuckDuckGo - gratis)
# =============================================================

def buscar_en_web(consulta: str, max_resultados: int = 5) -> str:
    try:
        from duckduckgo_search import DDGS
    except ImportError:
        return "Error: Instala duckduckgo-search con: pip install duckduckgo-search"

    for intento in range(3):
        try:
            with DDGS() as ddgs:
                resultados = list(ddgs.text(consulta, max_results=max_resultados, region="wt-wt"))
            if resultados:
                texto = f"Resultados de búsqueda para '{consulta}':\n\n"
                for i, r in enumerate(resultados, 1):
                    texto += f"{i}. **{r['title']}**\n"
                    texto += f"   {r['body']}\n"
                    texto += f"   Fuente: {r['href']}\n\n"
                return texto
        except Exception:
            if intento < 2:
                time.sleep(2)
                continue

    # Respaldo: buscar noticias
    try:
        with DDGS() as ddgs:
            resultados = list(ddgs.news(consulta, max_results=max_resultados))
        if resultados:
            texto = f"Noticias encontradas para '{consulta}':\n\n"
            for i, r in enumerate(resultados, 1):
                texto += f"{i}. **{r['title']}**\n"
                texto += f"   {r['body']}\n"
                texto += f"   Fuente: {r['source']}\n"
                texto += f"   Fecha: {r.get('date', 'N/A')}\n\n"
            return texto
    except Exception:
        pass

    return f"No se pudieron obtener resultados para '{consulta}'. Intenta de nuevo en unos minutos."


# =============================================================
# 2. 📧 ENVIAR CORREO (Gmail SMTP)
# =============================================================

def enviar_correo(destinatario: str, asunto: str, cuerpo: str) -> str:
    email_address = os.getenv("EMAIL_ADDRESS")
    email_password = os.getenv("EMAIL_PASSWORD")

    if not email_address or not email_password:
        return ("Error: Configura EMAIL_ADDRESS y EMAIL_PASSWORD en tu archivo .env. "
                "Necesitas una contraseña de aplicación de Google.")

    try:
        mensaje = MIMEMultipart()
        mensaje["From"] = email_address
        mensaje["To"] = destinatario
        mensaje["Subject"] = asunto
        mensaje.attach(MIMEText(cuerpo, "plain"))

        with smtplib.SMTP("smtp.gmail.com", 587) as servidor:
            servidor.starttls()
            servidor.login(email_address, email_password)
            servidor.send_message(mensaje)

        return f"Correo enviado exitosamente a {destinatario} con asunto '{asunto}'."

    except smtplib.SMTPAuthenticationError:
        return "Error de autenticación. Verifica tu EMAIL_PASSWORD (necesitas contraseña de app de Google)."
    except Exception as e:
        return f"Error al enviar correo: {str(e)}"


# =============================================================
# 3. 📄 LEER PDFs
# =============================================================

def leer_pdf(ruta_archivo: str) -> str:
    try:
        from PyPDF2 import PdfReader
    except ImportError:
        return "Error: Instala PyPDF2 con: pip install PyPDF2"

    try:
        ruta = Path(ruta_archivo)
        if not ruta.exists():
            return f"Error: No se encontró el archivo '{ruta_archivo}'."
        if ruta.suffix.lower() != ".pdf":
            return f"Error: '{ruta_archivo}' no es un PDF."

        reader = PdfReader(str(ruta))
        texto_completo = ""
        for i, pagina in enumerate(reader.pages, 1):
            texto = pagina.extract_text()
            if texto:
                texto_completo += f"\n--- Página {i} ---\n{texto}"

        if not texto_completo.strip():
            return "El PDF no contiene texto extraíble (puede ser escaneado)."

        if len(texto_completo) > 4000:
            texto_completo = texto_completo[:4000] + "\n\n[... texto truncado ...]"

        return f"Contenido de '{ruta.name}' ({len(reader.pages)} páginas):\n{texto_completo}"
    except Exception as e:
        return f"Error al leer PDF: {str(e)}"


# =============================================================
# 4. 📊 LEER EXCEL/CSV
# =============================================================

def leer_excel_csv(ruta_archivo: str, max_filas: int = 20) -> str:
    try:
        ruta = Path(ruta_archivo)
        if not ruta.exists():
            return f"Error: No se encontró '{ruta_archivo}'."

        extension = ruta.suffix.lower()

        if extension == ".csv":
            filas = []
            with open(ruta, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                for i, fila in enumerate(reader):
                    if i >= max_filas + 1:
                        break
                    filas.append(fila)
            if not filas:
                return "El archivo CSV está vacío."
            encabezados = filas[0]
            datos = filas[1:]
            total_filas = sum(1 for _ in open(ruta, "r", encoding="utf-8-sig")) - 1

        elif extension in [".xlsx", ".xls"]:
            try:
                from openpyxl import load_workbook
            except ImportError:
                return "Error: Instala openpyxl con: pip install openpyxl"
            wb = load_workbook(str(ruta), read_only=True)
            ws = wb.active
            filas = []
            for i, fila in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_filas + 1:
                    break
                filas.append([str(c) if c is not None else "" for c in fila])
            wb.close()
            if not filas:
                return "El archivo Excel está vacío."
            encabezados = filas[0]
            datos = filas[1:]
            total_filas = ws.max_row - 1 if ws.max_row else 0
        else:
            return f"Error: Formato no soportado '{extension}'. Usa .csv, .xlsx o .xls"

        resultado = f"Archivo: {ruta.name}\n"
        resultado += f"Columnas: {', '.join(encabezados)}\n"
        resultado += f"Total de filas: {total_filas}\n"
        resultado += f"Mostrando primeras {min(len(datos), max_filas)} filas:\n\n"
        resultado += " | ".join(encabezados) + "\n"
        resultado += "-" * (len(resultado.split(chr(10))[-2])) + "\n"
        for fila in datos:
            resultado += " | ".join(fila) + "\n"
        if total_filas > max_filas:
            resultado += f"\n[... {total_filas - max_filas} filas más ...]"

        return resultado
    except Exception as e:
        return f"Error al leer archivo: {str(e)}"


# =============================================================
# 5. 🗄️ BASE DE DATOS SQLite
# =============================================================

DB_PATH = DB_PATH = r"C:\Users\omurillo\.gemini\antigravity\scratch\claude_agent\agente_datos.db"

def inicializar_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS contactos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL, email TEXT, telefono TEXT,
            notas TEXT, creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tareas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT NOT NULL, descripcion TEXT,
            estado TEXT DEFAULT 'pendiente', prioridad TEXT DEFAULT 'media',
            creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("SELECT COUNT(*) FROM contactos")
    if cursor.fetchone()[0] == 0:
        cursor.executemany(
            "INSERT INTO contactos (nombre, email, telefono, notas) VALUES (?, ?, ?, ?)",
            [("Ana García", "ana@ejemplo.com", "33-1234-5678", "Cliente principal"),
             ("Carlos López", "carlos@ejemplo.com", "81-9876-5432", "Proveedor"),
             ("María Rodríguez", "maria@ejemplo.com", "55-5555-1234", "Socia de proyecto")]
        )
    cursor.execute("SELECT COUNT(*) FROM tareas")
    if cursor.fetchone()[0] == 0:
        cursor.executemany(
            "INSERT INTO tareas (titulo, descripcion, estado, prioridad) VALUES (?, ?, ?, ?)",
            [("Revisar propuesta", "Revisar la propuesta del cliente", "pendiente", "alta"),
             ("Enviar reporte", "Reporte mensual de ventas", "pendiente", "media"),
             ("Llamar a proveedor", "Negociar precios Q3", "completada", "baja")]
        )
    conn.commit()
    conn.close()

def consultar_base_datos(consulta_sql: str) -> str:
    consulta_upper = consulta_sql.strip().upper()
    if any(p in consulta_upper for p in ["DROP", "DELETE", "ALTER", "TRUNCATE"]):
        return "Error: Operaciones DROP, DELETE, ALTER y TRUNCATE no permitidas por seguridad."
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute(consulta_sql)
        if consulta_upper.startswith("SELECT"):
            columnas = [desc[0] for desc in cursor.description] if cursor.description else []
            filas = cursor.fetchall()
            if not filas:
                resultado = "La consulta no devolvió resultados."
            else:
                resultado = f"Columnas: {', '.join(columnas)}\nResultados ({len(filas)} filas):\n\n"
                for fila in filas:
                    resultado += " | ".join(str(v) for v in fila) + "\n"
        else:
            conn.commit()
            resultado = f"Consulta ejecutada. Filas afectadas: {cursor.rowcount}"
        conn.close()
        return resultado
    except sqlite3.Error as e:
        return f"Error SQL: {str(e)}"
    except Exception as e:
        return f"Error: {str(e)}"


# =============================================================
# 6. 💾 GUARDAR ARCHIVOS
# =============================================================

def guardar_archivo(nombre_archivo: str, contenido: str) -> str:
    try:
        ruta = Path(nombre_archivo)
        extensiones_permitidas = [".txt", ".md", ".csv", ".json", ".html"]
        if ruta.suffix.lower() not in extensiones_permitidas:
            return f"Error: Solo se permiten archivos {', '.join(extensiones_permitidas)}"
        with open(ruta, "w", encoding="utf-8") as f:
            f.write(contenido)
        return f"Archivo '{nombre_archivo}' guardado exitosamente ({len(contenido)} caracteres)."
    except Exception as e:
        return f"Error al guardar archivo: {str(e)}"


# =============================================================
# 7. 🌐 TRADUCTOR DE IDIOMAS
# =============================================================

def traducir_texto(texto: str, idioma_destino: str, idioma_origen: str = "auto") -> str:
    try:
        client = Anthropic()
        prompt = f"""Traduce el siguiente texto al {idioma_destino}.
Si el idioma de origen es "auto", detecta automáticamente el idioma.
Idioma de origen: {idioma_origen}

REGLAS:
- Devuelve SOLO la traducción, sin explicaciones
- Mantén el formato original
- No traduzcas nombres propios

Texto a traducir:
{texto}"""

        respuesta = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}]
        )
        traduccion = "".join(b.text for b in respuesta.content if hasattr(b, "text"))
        idioma_info = " (detectado automáticamente)" if idioma_origen == "auto" else ""
        return (f"Traducción al {idioma_destino}{idioma_info}:\n\n{traduccion}\n\n"
                f"--- Original ({len(texto)} chars) → Traducción ({len(traduccion)} chars) ---")
    except Exception as e:
        return f"Error al traducir: {str(e)}"


# =============================================================
# 8. 📈 GENERAR GRÁFICAS
# =============================================================

def generar_grafica(
    ruta_csv: str, tipo_grafica: str = "barras",
    columna_x: str = "", columna_y: str = "",
    titulo: str = "Gráfica", nombre_archivo: str = "grafica.png"
) -> str:
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.ticker as ticker
    except ImportError:
        return "Error: Instala matplotlib con: pip install matplotlib"

    try:
        ruta = Path(ruta_csv)
        if not ruta.exists():
            return f"Error: No se encontró '{ruta_csv}'."

        # Leer datos
        datos = []
        encabezados = []
        extension = ruta.suffix.lower()

        if extension == ".csv":
            with open(ruta, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                encabezados = next(reader)
                for fila in reader:
                    datos.append(fila)
        elif extension in [".xlsx", ".xls"]:
            from openpyxl import load_workbook
            wb = load_workbook(str(ruta), read_only=True)
            ws = wb.active
            filas = list(ws.iter_rows(values_only=True))
            wb.close()
            if filas:
                encabezados = [str(c) if c else f"Col{i}" for i, c in enumerate(filas[0])]
                datos = [[str(c) if c else "" for c in fila] for fila in filas[1:]]
        else:
            return f"Error: Usa .csv o .xlsx"

        if not datos or not encabezados:
            return "Error: Archivo vacío."

        # Auto-detectar columnas
        if not columna_x:
            columna_x = encabezados[0]
        if not columna_y:
            for i, h in enumerate(encabezados[1:], 1):
                try:
                    float(datos[0][i].replace(",", "").replace("$", ""))
                    columna_y = h
                    break
                except (ValueError, IndexError):
                    continue
            if not columna_y:
                columna_y = encabezados[1] if len(encabezados) > 1 else encabezados[0]

        # Encontrar índices
        col_x_idx = col_y_idx = None
        for i, h in enumerate(encabezados):
            if h.strip().lower() == columna_x.strip().lower():
                col_x_idx = i
            if h.strip().lower() == columna_y.strip().lower():
                col_y_idx = i

        if col_x_idx is None:
            return f"Error: Columna '{columna_x}' no encontrada. Disponibles: {', '.join(encabezados)}"
        if col_y_idx is None:
            return f"Error: Columna '{columna_y}' no encontrada. Disponibles: {', '.join(encabezados)}"

        # Agrupar datos
        agrupado = {}
        for fila in datos:
            try:
                x_val = fila[col_x_idx].strip()
                y_val = float(fila[col_y_idx].strip().replace(",", "").replace("$", ""))
                agrupado[x_val] = agrupado.get(x_val, 0) + y_val
            except (ValueError, IndexError):
                continue

        if not agrupado:
            return "Error: No se pudieron extraer datos numéricos."

        etiquetas_x = list(agrupado.keys())
        valores_y = list(agrupado.values())

        # Crear gráfica
        plt.style.use("seaborn-v0_8-whitegrid")
        fig, ax = plt.subplots(figsize=(10, 6))
        colores = ["#378ADD", "#1D9E75", "#D85A30", "#D4537E", "#BA7517",
                    "#534AB7", "#639922", "#E24B4A", "#888780"]
        tipo = tipo_grafica.lower().strip()

        if tipo in ["barras", "bar"]:
            barras = ax.bar(etiquetas_x, valores_y, color=colores[:len(etiquetas_x)], edgecolor="white")
            for b, v in zip(barras, valores_y):
                ax.text(b.get_x() + b.get_width()/2., b.get_height(), f'{v:,.0f}',
                       ha='center', va='bottom', fontsize=9, fontweight='bold')
        elif tipo in ["lineas", "linea", "line"]:
            ax.plot(etiquetas_x, valores_y, color=colores[0], marker="o", linewidth=2,
                   markersize=8, markerfacecolor="white", markeredgecolor=colores[0], markeredgewidth=2)
            for x, y in zip(etiquetas_x, valores_y):
                ax.annotate(f'{y:,.0f}', (x, y), textcoords="offset points", xytext=(0, 12),
                           ha='center', fontsize=9, fontweight='bold')
        elif tipo in ["pastel", "pie"]:
            ax.pie(valores_y, labels=etiquetas_x, colors=colores[:len(etiquetas_x)],
                  autopct='%1.1f%%', startangle=90, textprops={'fontsize': 11})
            ax.axis('equal')
        elif tipo in ["barras_horizontales", "horizontal", "barh"]:
            barras = ax.barh(etiquetas_x, valores_y, color=colores[:len(etiquetas_x)], edgecolor="white")
            for b, v in zip(barras, valores_y):
                ax.text(b.get_width() + max(valores_y)*0.01, b.get_y() + b.get_height()/2.,
                       f'{v:,.0f}', ha='left', va='center', fontsize=9, fontweight='bold')
        elif tipo in ["dispersión", "dispersion", "scatter"]:
            ax.scatter(range(len(valores_y)), valores_y, c=colores[0], s=100, edgecolors="white", zorder=5)
            ax.set_xticks(range(len(etiquetas_x)))
            ax.set_xticklabels(etiquetas_x)
        else:
            return f"Error: Tipo '{tipo_grafica}' no soportado. Usa: barras, lineas, pastel, barras_horizontales, dispersión"

        ax.set_title(titulo, fontsize=16, fontweight="bold", pad=15)
        if tipo not in ["pastel", "pie"]:
            ax.set_xlabel(columna_x, fontsize=12, labelpad=10)
            ax.set_ylabel(columna_y, fontsize=12, labelpad=10)
            ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, p: f'{x:,.0f}'))
            plt.xticks(rotation=45 if len(etiquetas_x) > 5 else 0, ha='right')

        plt.tight_layout()
        if Path(nombre_archivo).suffix.lower() not in [".png", ".jpg", ".pdf", ".svg"]:
            nombre_archivo += ".png"
        plt.savefig(nombre_archivo, dpi=150, bbox_inches='tight', facecolor='white')
        plt.close()

        return (f"Gráfica generada exitosamente:\n"
                f"  Archivo: {nombre_archivo}\n  Tipo: {tipo_grafica}\n"
                f"  Datos: {columna_x} vs {columna_y}\n"
                f"  Valores: {', '.join(f'{e}={v:,.0f}' for e, v in zip(etiquetas_x, valores_y))}\n\n"
                f"Ábrela con cualquier visor de imágenes.")
    except Exception as e:
        return f"Error al generar gráfica: {str(e)}"


# =============================================================
# 9. 📝 RESUMIDOR DE TEXTOS
# =============================================================

def resumir_texto(texto: str, longitud: str = "medio", idioma: str = "español") -> str:
    """Genera un resumen de un texto largo usando Claude."""
    try:
        client = Anthropic()

        instrucciones_longitud = {
            "corto": "Resumen en máximo 3 oraciones.",
            "medio": "Resumen en 1-2 párrafos cortos con los puntos clave.",
            "largo": "Resumen detallado con todos los puntos importantes, organizado por temas.",
            "puntos": "Lista de los 5-7 puntos más importantes como viñetas."
        }

        instruccion = instrucciones_longitud.get(longitud.lower(), instrucciones_longitud["medio"])

        prompt = f"""Genera un resumen del siguiente texto.
Idioma del resumen: {idioma}
Formato: {instruccion}

REGLAS:
- Captura las ideas principales y datos clave
- No agregues información que no esté en el texto original
- Usa un tono claro y profesional

Texto a resumir:
{texto}"""

        respuesta = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}]
        )
        resumen = "".join(b.text for b in respuesta.content if hasattr(b, "text"))

        return (f"Resumen ({longitud}):\n\n{resumen}\n\n"
                f"--- Texto original: {len(texto)} caracteres → Resumen: {len(resumen)} caracteres "
                f"(reducción del {100 - int(len(resumen)/max(len(texto),1)*100)}%) ---")
    except Exception as e:
        return f"Error al resumir: {str(e)}"


# =============================================================
# 10. 🌤️ CLIMA EN TIEMPO REAL (OpenWeatherMap)
# =============================================================

def obtener_clima(ciudad: str, pais: str = "MX") -> str:
    """Obtiene el clima real de una ciudad usando OpenWeatherMap."""
    api_key = os.getenv("OPENWEATHER_API_KEY")

    if not api_key:
        # Respaldo sin API: datos informativos
        return (f"Para obtener el clima real de {ciudad}, configura OPENWEATHER_API_KEY en tu .env. "
                f"Obtén una gratis en: https://openweathermap.org/api (plan gratuito).\n\n"
                f"Sin la API key, no puedo consultar el clima en tiempo real.")

    try:
        import requests
    except ImportError:
        return "Error: Instala requests con: pip install requests"

    try:
        url = "https://api.openweathermap.org/data/2.5/weather"
        params = {
            "q": f"{ciudad},{pais}",
            "appid": api_key,
            "units": "metric",
            "lang": "es"
        }
        resp = requests.get(url, params=params, timeout=10)
        data = resp.json()

        if resp.status_code != 200:
            return f"Error: {data.get('message', 'Ciudad no encontrada')}. Verifica el nombre de la ciudad."

        temp = data["main"]["temp"]
        sensacion = data["main"]["feels_like"]
        humedad = data["main"]["humidity"]
        descripcion = data["weather"][0]["description"].capitalize()
        viento = data["wind"]["speed"]
        temp_min = data["main"]["temp_min"]
        temp_max = data["main"]["temp_max"]
        nombre_ciudad = data["name"]
        pais_code = data["sys"]["country"]

        return (
            f"Clima actual en {nombre_ciudad}, {pais_code}:\n\n"
            f"  🌡️ Temperatura: {temp:.1f}°C (sensación de {sensacion:.1f}°C)\n"
            f"  📊 Mín: {temp_min:.1f}°C | Máx: {temp_max:.1f}°C\n"
            f"  🌤️ Condición: {descripcion}\n"
            f"  💧 Humedad: {humedad}%\n"
            f"  💨 Viento: {viento} m/s\n"
        )
    except requests.Timeout:
        return "Error: La consulta tardó demasiado. Intenta de nuevo."
    except Exception as e:
        return f"Error al consultar clima: {str(e)}"


# =============================================================
# DEFINICIÓN DE HERRAMIENTAS PARA CLAUDE
# =============================================================

HERRAMIENTAS = [
    {
        "name": "buscar_en_web",
        "description": "Busca información actualizada en internet. Úsala para noticias, datos recientes o cualquier tema que requiera información en tiempo real.",
        "input_schema": {
            "type": "object",
            "properties": {
                "consulta": {"type": "string", "description": "Lo que quieres buscar"},
                "max_resultados": {"type": "integer", "description": "Máximo de resultados (1-10)", "default": 5}
            },
            "required": ["consulta"]
        }
    },
    {
        "name": "enviar_correo",
        "description": "Envía un correo electrónico. Úsala cuando el usuario pida enviar un email.",
        "input_schema": {
            "type": "object",
            "properties": {
                "destinatario": {"type": "string", "description": "Dirección de correo del destinatario"},
                "asunto": {"type": "string", "description": "Asunto del correo"},
                "cuerpo": {"type": "string", "description": "Contenido del correo"}
            },
            "required": ["destinatario", "asunto", "cuerpo"]
        }
    },
    {
        "name": "leer_pdf",
        "description": "Lee y extrae texto de un PDF. Necesita la ruta completa al archivo.",
        "input_schema": {
            "type": "object",
            "properties": {
                "ruta_archivo": {"type": "string", "description": "Ruta completa al PDF"}
            },
            "required": ["ruta_archivo"]
        }
    },
    {
        "name": "leer_excel_csv",
        "description": "Lee archivos Excel (.xlsx) o CSV (.csv) y muestra su contenido.",
        "input_schema": {
            "type": "object",
            "properties": {
                "ruta_archivo": {"type": "string", "description": "Ruta completa al archivo"},
                "max_filas": {"type": "integer", "description": "Máximo de filas a mostrar", "default": 20}
            },
            "required": ["ruta_archivo"]
        }
    },
    {
        "name": "consultar_base_datos",
        "description": (
            "Ejecuta consultas SQL. Tablas: 'contactos' (id, nombre, email, telefono, notas, creado_en) "
            "y 'tareas' (id, titulo, descripcion, estado, prioridad, creado_en). Soporta SELECT, INSERT, UPDATE."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "consulta_sql": {"type": "string", "description": "Consulta SQL a ejecutar"}
            },
            "required": ["consulta_sql"]
        }
    },
    {
        "name": "guardar_archivo",
        "description": "Guarda texto en un archivo (.txt, .md, .csv, .json, .html).",
        "input_schema": {
            "type": "object",
            "properties": {
                "nombre_archivo": {"type": "string", "description": "Nombre del archivo con extensión"},
                "contenido": {"type": "string", "description": "Texto a guardar"}
            },
            "required": ["nombre_archivo", "contenido"]
        }
    },
    {
        "name": "traducir_texto",
        "description": "Traduce texto entre idiomas. Soporta español, inglés, francés, portugués, alemán, italiano, japonés, chino, coreano y más.",
        "input_schema": {
            "type": "object",
            "properties": {
                "texto": {"type": "string", "description": "Texto a traducir"},
                "idioma_destino": {"type": "string", "description": "Idioma destino (ej: inglés, francés)"},
                "idioma_origen": {"type": "string", "description": "Idioma origen ('auto' para detectar)", "default": "auto"}
            },
            "required": ["texto", "idioma_destino"]
        }
    },
    {
        "name": "generar_grafica",
        "description": "Genera gráficas (barras, lineas, pastel, barras_horizontales, dispersión) a partir de CSV o Excel.",
        "input_schema": {
            "type": "object",
            "properties": {
                "ruta_csv": {"type": "string", "description": "Ruta al archivo CSV o Excel"},
                "tipo_grafica": {"type": "string", "description": "Tipo: barras, lineas, pastel, barras_horizontales, dispersión", "default": "barras"},
                "columna_x": {"type": "string", "description": "Columna eje X (auto si se omite)", "default": ""},
                "columna_y": {"type": "string", "description": "Columna eje Y (auto si se omite)", "default": ""},
                "titulo": {"type": "string", "description": "Título de la gráfica", "default": "Gráfica"},
                "nombre_archivo": {"type": "string", "description": "Archivo de salida (ej: ventas.png)", "default": "grafica.png"}
            },
            "required": ["ruta_csv"]
        }
    },
    {
        "name": "resumir_texto",
        "description": "Resume textos largos en versiones cortas. Tipos de resumen: corto (3 oraciones), medio (1-2 párrafos), largo (detallado), puntos (viñetas). Úsala cuando pidan resumir, sintetizar o extraer puntos clave.",
        "input_schema": {
            "type": "object",
            "properties": {
                "texto": {"type": "string", "description": "Texto largo a resumir"},
                "longitud": {"type": "string", "description": "Tipo: corto, medio, largo, puntos", "default": "medio"},
                "idioma": {"type": "string", "description": "Idioma del resumen", "default": "español"}
            },
            "required": ["texto"]
        }
    },
    {
        "name": "obtener_clima",
        "description": "Consulta el clima actual de una ciudad en tiempo real. Muestra temperatura, humedad, viento y condiciones. Úsala cuando pregunten por el clima o temperatura.",
        "input_schema": {
            "type": "object",
            "properties": {
                "ciudad": {"type": "string", "description": "Nombre de la ciudad (ej: Guadalajara, CDMX)"},
                "pais": {"type": "string", "description": "Código de país (ej: MX, US, ES)", "default": "MX"}
            },
            "required": ["ciudad"]
        }
    }
]

# Mapeo de funciones
FUNCIONES_DISPONIBLES = {
    "buscar_en_web": lambda a: buscar_en_web(a["consulta"], a.get("max_resultados", 5)),
    "enviar_correo": lambda a: enviar_correo(a["destinatario"], a["asunto"], a["cuerpo"]),
    "leer_pdf": lambda a: leer_pdf(a["ruta_archivo"]),
    "leer_excel_csv": lambda a: leer_excel_csv(a["ruta_archivo"], a.get("max_filas", 20)),
    "consultar_base_datos": lambda a: consultar_base_datos(a["consulta_sql"]),
    "guardar_archivo": lambda a: guardar_archivo(a["nombre_archivo"], a["contenido"]),
    "traducir_texto": lambda a: traducir_texto(a["texto"], a["idioma_destino"], a.get("idioma_origen", "auto")),
    "generar_grafica": lambda a: generar_grafica(a["ruta_csv"], a.get("tipo_grafica", "barras"), a.get("columna_x", ""), a.get("columna_y", ""), a.get("titulo", "Gráfica"), a.get("nombre_archivo", "grafica.png")),
    "resumir_texto": lambda a: resumir_texto(a["texto"], a.get("longitud", "medio"), a.get("idioma", "español")),
    "obtener_clima": lambda a: obtener_clima(a["ciudad"], a.get("pais", "MX")),
}


# =============================================================
# SYSTEM PROMPT
# =============================================================

SYSTEM_PROMPT = """
Eres un asistente personal inteligente y completo que habla español.
Tu nombre es "Agente Pro" y tu objetivo es ayudar al usuario con
cualquier tarea usando las herramientas disponibles.

## TUS HERRAMIENTAS:

1. **buscar_en_web** - Buscar información actualizada en internet
2. **enviar_correo** - Enviar correos electrónicos
3. **leer_pdf** - Leer y extraer texto de archivos PDF
4. **leer_excel_csv** - Leer hojas de cálculo Excel y archivos CSV
5. **consultar_base_datos** - Consultar y modificar la base de datos
   (tablas: contactos y tareas)
6. **guardar_archivo** - Guardar resultados en archivos de texto
7. **traducir_texto** - Traducir texto entre idiomas
8. **generar_grafica** - Crear gráficas a partir de datos CSV/Excel
9. **resumir_texto** - Resumir textos largos en versiones cortas
10. **obtener_clima** - Consultar el clima actual de cualquier ciudad

## REGLAS:

1. Siempre responde en español de forma amigable y profesional
2. Usa las herramientas cuando sea necesario - no inventes datos
3. Para búsquedas web, resume los resultados de forma clara
4. Antes de enviar un correo, confirma con el usuario el contenido
5. Si no puedes hacer algo, explica por qué y sugiere alternativas
6. Cuando uses la base de datos, explica qué consulta estás haciendo
7. Si el usuario te pide guardar algo, usa un nombre de archivo descriptivo
8. Puedes combinar múltiples herramientas para tareas complejas
9. Para traducir, detecta automáticamente el idioma si no se especifica
10. Para gráficas, elige el tipo más adecuado si no se especifica

## EJEMPLOS DE TAREAS COMPLEJAS:

- "Busca noticias sobre IA y guárdalas en un archivo"
  → buscar_en_web + guardar_archivo
- "Lee mi PDF y envíame un resumen por correo"
  → leer_pdf + resumir_texto + enviar_correo
- "Genera una gráfica de ventas y traduce el resumen al inglés"
  → leer_excel_csv + generar_grafica + traducir_texto
- "¿Qué clima hace en Guadalajara?"
  → obtener_clima
- "Resume este texto en puntos clave y guárdalo"
  → resumir_texto + guardar_archivo
"""


# =============================================================
# LOOP PRINCIPAL
# =============================================================

def ejecutar_herramienta(nombre: str, argumentos: dict) -> str:
    if nombre in FUNCIONES_DISPONIBLES:
        return FUNCIONES_DISPONIBLES[nombre](argumentos)
    return f"Herramienta '{nombre}' no encontrada."


def agente_completo():
    inicializar_db()
    client = Anthropic()
    mensajes = []

    print("")
    print("=" * 58)
    print("  🤖 AGENTE PRO ULTIMATE - Tu Asistente Inteligente")
    print("=" * 58)
    print("")
    print("  Herramientas disponibles (10):")
    print("  🔍 Búsqueda web      📧 Enviar correos")
    print("  📄 Leer PDFs          📊 Leer Excel/CSV")
    print("  🗄️  Base de datos      💾 Guardar archivos")
    print("  🌐 Traductor          📈 Generar gráficas")
    print("  📝 Resumidor          🌤️  Clima en tiempo real")
    print("")
    print("  Ejemplos:")
    print('  • "Busca noticias sobre IA y guárdalas"')
    print('  • "¿Qué clima hace en Guadalajara?"')
    print('  • "Lee mi PDF y resume los puntos clave"')
    print('  • "Traduce al inglés: Hola, ¿cómo estás?"')
    print('  • "Genera una gráfica de barras con ventas.csv"')
    print('  • "¿Qué tareas pendientes tengo?"')
    print("")
    print('  Escribe "salir" para terminar.')
    print("=" * 58)

    while True:
        pregunta = input("\n🧑 Tú: ").strip()
        if pregunta.lower() in ["salir", "exit", "quit"]:
            print("\n👋 ¡Hasta luego! Tus datos están guardados.")
            break
        if not pregunta:
            continue

        mensajes.append({"role": "user", "content": pregunta})

        while True:
            try:
                respuesta = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=2048,
                    system=SYSTEM_PROMPT,
                    tools=HERRAMIENTAS,
                    messages=mensajes,
                )
            except Exception as e:
                print(f"\n❌ Error al conectar con Claude: {str(e)}")
                mensajes.pop()
                break

            if respuesta.stop_reason == "tool_use":
                mensajes.append({"role": "assistant", "content": respuesta.content})
                resultados = []
                for bloque in respuesta.content:
                    if bloque.type == "tool_use":
                        print(f"  🔧 Usando: {bloque.name}")
                        resultado = ejecutar_herramienta(bloque.name, bloque.input)
                        print(f"  ✅ Listo")
                        resultados.append({
                            "type": "tool_result",
                            "tool_use_id": bloque.id,
                            "content": resultado,
                        })
                mensajes.append({"role": "user", "content": resultados})
            else:
                texto = "".join(b.text for b in respuesta.content if hasattr(b, "text"))
                print(f"\n🤖 Agente Pro: {texto}")
                mensajes.append({"role": "assistant", "content": respuesta.content})
                break


if __name__ == "__main__":
    agente_completo()
