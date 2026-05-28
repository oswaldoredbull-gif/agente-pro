"""
=============================================================
🤖 SISTEMA MULTI-AGENTE - Interfaz Web
=============================================================

Interfaz visual para ver cómo el Investigador y el Escritor
colaboran en tiempo real.

INSTALACIÓN:
  pip install streamlit anthropic python-dotenv duckduckgo-search PyPDF2 openpyxl

EJECUCIÓN:
  streamlit run multi_agente_web.py

=============================================================
"""

import streamlit as st
import os
import csv
import time
import json
from pathlib import Path
from datetime import datetime

from dotenv import load_dotenv
from anthropic import Anthropic

load_dotenv()

st.set_page_config(
    page_title="Multi-Agente Pro",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700&family=JetBrains+Mono:wght@400;500&display=swap');
    .stApp {
        background: linear-gradient(135deg, #0f0f23 0%, #1a1a3e 50%, #0d0d2b 100%);
    }
    .agent-card {
        border-radius: 16px;
        padding: 20px;
        margin: 8px 0;
        border: 1px solid rgba(255,255,255,0.06);
        backdrop-filter: blur(10px);
    }
    .coordinator { background: linear-gradient(135deg, rgba(83,74,183,0.15), rgba(83,74,183,0.05)); border-left: 3px solid #7F77DD; }
    .researcher { background: linear-gradient(135deg, rgba(29,158,117,0.15), rgba(29,158,117,0.05)); border-left: 3px solid #5DCAA5; }
    .writer { background: linear-gradient(135deg, rgba(216,90,48,0.15), rgba(216,90,48,0.05)); border-left: 3px solid #F0997B; }
    .agent-name {
        font-size: 14px;
        font-weight: 600;
        margin-bottom: 8px;
        display: flex;
        align-items: center;
        gap: 8px;
    }
    .coord-color { color: #AFA9EC; }
    .research-color { color: #5DCAA5; }
    .writer-color { color: #F0997B; }
    .agent-content {
        font-size: 14px;
        color: #cbd5e1;
        line-height: 1.7;
        white-space: pre-wrap;
    }
    .tool-pill {
        display: inline-flex;
        align-items: center;
        gap: 4px;
        background: rgba(16,185,129,0.12);
        color: #34d399;
        padding: 3px 10px;
        border-radius: 20px;
        font-size: 12px;
        font-family: 'JetBrains Mono', monospace;
        margin: 3px 2px;
        border: 1px solid rgba(16,185,129,0.15);
    }
    .step-badge {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 28px;
        height: 28px;
        border-radius: 50%;
        font-size: 13px;
        font-weight: 600;
        margin-right: 8px;
    }
    .step-active { background: rgba(96,165,250,0.2); color: #60a5fa; border: 1px solid rgba(96,165,250,0.3); }
    .step-done { background: rgba(16,185,129,0.2); color: #34d399; border: 1px solid rgba(16,185,129,0.3); }
    .step-waiting { background: rgba(255,255,255,0.05); color: #475569; border: 1px solid rgba(255,255,255,0.08); }
    .result-card {
        background: rgba(255,255,255,0.04);
        border: 1px solid rgba(255,255,255,0.08);
        border-radius: 16px;
        padding: 24px;
        margin: 16px 0;
        color: #e2e8f0;
        font-size: 15px;
        line-height: 1.8;
        white-space: pre-wrap;
    }
    .header-title {
        font-size: 28px;
        font-weight: 700;
        background: linear-gradient(135deg, #60a5fa, #a78bfa, #f472b6);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
    }
    .header-sub {
        font-size: 14px;
        color: #64748b;
        text-align: center;
        margin-bottom: 24px;
    }
    .user-msg {
        background: linear-gradient(135deg, #2563eb, #1d4ed8);
        color: white;
        padding: 12px 18px;
        border-radius: 16px 16px 4px 16px;
        margin: 8px 0;
        max-width: 80%;
        margin-left: auto;
        font-size: 14px;
    }
    .flow-arrow {
        text-align: center;
        color: #475569;
        font-size: 20px;
        margin: 4px 0;
    }
    [data-testid="stSidebar"] {
        background: rgba(15,15,35,0.95) !important;
        border-right: 1px solid rgba(255,255,255,0.06);
    }
    .example-btn {
        background: rgba(255,255,255,0.04);
        border: 1px solid rgba(255,255,255,0.08);
        border-radius: 10px;
        padding: 10px 14px;
        color: #94a3b8;
        font-size: 13px;
        margin: 4px 0;
        cursor: pointer;
        width: 100%;
        text-align: left;
        transition: all 0.2s;
    }
    .example-btn:hover {
        background: rgba(255,255,255,0.08);
        color: #e2e8f0;
    }
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .stTextInput > div > div > input {
        background: rgba(255,255,255,0.06) !important;
        border: 1px solid rgba(255,255,255,0.1) !important;
        border-radius: 14px !important;
        color: #e2e8f0 !important;
        padding: 14px 18px !important;
        font-size: 15px !important;
    }
    .stTextInput > div > div > input:focus {
        border-color: rgba(96,165,250,0.5) !important;
    }
    .stTextInput > div > div > input::placeholder { color: #475569 !important; }
</style>
""", unsafe_allow_html=True)


# =============================================================
# HERRAMIENTAS
# =============================================================

def buscar_en_web(consulta, max_resultados=5):
    try:
        from duckduckgo_search import DDGS
    except ImportError:
        return "Error: pip install duckduckgo-search"
    for intento in range(3):
        try:
            with DDGS() as ddgs:
                resultados = list(ddgs.text(consulta, max_results=max_resultados, region="wt-wt"))
            if resultados:
                texto = ""
                for i, r in enumerate(resultados, 1):
                    texto += f"{i}. {r['title']}\n   {r['body']}\n   Fuente: {r['href']}\n\n"
                return texto
        except Exception:
            if intento < 2: time.sleep(2)
    return f"Sin resultados para '{consulta}'."

def leer_pdf(ruta_archivo):
    try:
        from PyPDF2 import PdfReader
        ruta = Path(ruta_archivo)
        if not ruta.exists(): return f"No se encontró '{ruta_archivo}'"
        reader = PdfReader(str(ruta))
        texto = ""
        for i, p in enumerate(reader.pages, 1):
            t = p.extract_text()
            if t: texto += f"\n--- Página {i} ---\n{t}"
        if len(texto) > 4000: texto = texto[:4000] + "\n[truncado]"
        return texto if texto.strip() else "PDF sin texto."
    except Exception as e:
        return f"Error: {str(e)}"

def leer_excel_csv(ruta_archivo, max_filas=20):
    try:
        ruta = Path(ruta_archivo)
        if not ruta.exists(): return f"No se encontró '{ruta_archivo}'"
        ext = ruta.suffix.lower()
        if ext == ".csv":
            filas = []
            with open(ruta, "r", encoding="utf-8-sig") as f:
                for i, fila in enumerate(csv.reader(f)):
                    if i >= max_filas + 1: break
                    filas.append(fila)
        elif ext in [".xlsx", ".xls"]:
            from openpyxl import load_workbook
            wb = load_workbook(str(ruta), read_only=True); ws = wb.active
            filas = [[str(c) if c else "" for c in r] for i, r in enumerate(ws.iter_rows(values_only=True)) if i < max_filas + 1]
            wb.close()
        else: return f"Formato no soportado"
        if not filas: return "Vacío"
        enc, dat = filas[0], filas[1:]
        res = f"Columnas: {', '.join(enc)}\n"
        for f in dat: res += " | ".join(f) + "\n"
        return res
    except Exception as e:
        return f"Error: {str(e)}"

def guardar_archivo(nombre, contenido):
    try:
        r = Path(nombre)
        if r.suffix.lower() not in [".txt",".md",".csv",".json",".html"]: return "Extensión no permitida."
        with open(r, "w", encoding="utf-8") as f: f.write(contenido)
        return f"Archivo '{nombre}' guardado ({len(contenido)} chars)."
    except Exception as e:
        return f"Error: {str(e)}"

def traducir_texto_simple(texto, idioma_destino):
    try:
        client = Anthropic()
        resp = client.messages.create(model="claude-sonnet-4-6", max_tokens=2048,
            messages=[{"role":"user","content":f"Traduce al {idioma_destino}. Solo la traducción:\n{texto}"}])
        return "".join(b.text for b in resp.content if hasattr(b,"text"))
    except Exception as e:
        return f"Error: {str(e)}"


# =============================================================
# CONFIGURACIÓN DE AGENTES
# =============================================================

PROMPT_COORDINADOR = """Eres el coordinador de un sistema multi-agente. Analiza la solicitud y crea un plan.
Agentes: INVESTIGADOR (busca web, lee PDFs/Excel) y ESCRITOR (redacta documentos, guarda archivos, traduce).
Responde SOLO con JSON válido sin backticks:
{"plan":"descripción","pasos":[{"agente":"investigador","tarea":"..."},{"agente":"escritor","tarea":"..."}]}"""

PROMPT_INVESTIGADOR = """Eres un investigador experto. Busca y organiza información exhaustivamente.
Incluye datos específicos, estadísticas y fuentes. Responde en español.
Formato: HALLAZGOS PRINCIPALES, DATOS Y ESTADÍSTICAS, FUENTES, CONTEXTO."""

PROMPT_ESCRITOR = """Eres un escritor profesional. Transforma datos en documentos pulidos y atractivos.
Escribe en español con estructura clara, introducción y conclusión.
Si te piden guardar, usa guardar_archivo."""

HERRAMIENTAS_INV = [
    {"name":"buscar_en_web","description":"Busca en internet.","input_schema":{"type":"object","properties":{"consulta":{"type":"string"},"max_resultados":{"type":"integer","default":5}},"required":["consulta"]}},
    {"name":"leer_pdf","description":"Lee PDF.","input_schema":{"type":"object","properties":{"ruta_archivo":{"type":"string"}},"required":["ruta_archivo"]}},
    {"name":"leer_excel_csv","description":"Lee Excel/CSV.","input_schema":{"type":"object","properties":{"ruta_archivo":{"type":"string"},"max_filas":{"type":"integer","default":20}},"required":["ruta_archivo"]}},
]

HERRAMIENTAS_ESC = [
    {"name":"guardar_archivo","description":"Guarda documento.","input_schema":{"type":"object","properties":{"nombre_archivo":{"type":"string"},"contenido":{"type":"string"}},"required":["nombre_archivo","contenido"]}},
    {"name":"traducir_texto","description":"Traduce texto.","input_schema":{"type":"object","properties":{"texto":{"type":"string"},"idioma_destino":{"type":"string"}},"required":["texto","idioma_destino"]}},
]

FUNCIONES_INV = {
    "buscar_en_web": lambda a: buscar_en_web(a["consulta"], a.get("max_resultados",5)),
    "leer_pdf": lambda a: leer_pdf(a["ruta_archivo"]),
    "leer_excel_csv": lambda a: leer_excel_csv(a["ruta_archivo"], a.get("max_filas",20)),
}

FUNCIONES_ESC = {
    "guardar_archivo": lambda a: guardar_archivo(a["nombre_archivo"], a["contenido"]),
    "traducir_texto": lambda a: traducir_texto_simple(a["texto"], a["idioma_destino"]),
}


# =============================================================
# FUNCIONES DE AGENTES
# =============================================================

def ejecutar_agente(nombre, emoji, system_prompt, herramientas, funciones, tarea, contexto=""):
    client = Anthropic()
    prompt = tarea
    if contexto:
        prompt = f"CONTEXTO PREVIO:\n{contexto}\n\nTAREA:\n{tarea}"

    mensajes = [{"role": "user", "content": prompt}]
    tools_usadas = []

    while True:
        try:
            resp = client.messages.create(
                model="claude-sonnet-4-6", max_tokens=2048,
                system=system_prompt, tools=herramientas, messages=mensajes)
        except Exception as e:
            return f"Error: {str(e)}", tools_usadas

        if resp.stop_reason == "tool_use":
            mensajes.append({"role": "assistant", "content": resp.content})
            resultados = []
            for b in resp.content:
                if b.type == "tool_use":
                    tools_usadas.append(b.name)
                    resultado = funciones[b.name](b.input) if b.name in funciones else "No disponible"
                    resultados.append({"type": "tool_result", "tool_use_id": b.id, "content": resultado})
            mensajes.append({"role": "user", "content": resultados})
        else:
            texto = "".join(b.text for b in resp.content if hasattr(b, "text"))
            return texto, tools_usadas


def planificar(solicitud):
    client = Anthropic()
    try:
        resp = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=500,
            system=PROMPT_COORDINADOR,
            messages=[{"role": "user", "content": solicitud}])
        texto = "".join(b.text for b in resp.content if hasattr(b, "text")).strip()
        if texto.startswith("```"): texto = texto.split("\n", 1)[1] if "\n" in texto else texto[3:]
        if texto.endswith("```"): texto = texto[:-3]
        return json.loads(texto.strip())
    except Exception:
        return {
            "plan": "Investigar y escribir",
            "pasos": [
                {"agente": "investigador", "tarea": f"Investiga: {solicitud}"},
                {"agente": "escritor", "tarea": f"Escribe un documento sobre: {solicitud}"}
            ]
        }


# =============================================================
# INTERFAZ
# =============================================================

if "history" not in st.session_state:
    st.session_state.history = []

# Sidebar
with st.sidebar:
    st.markdown('<div style="font-size:22px;font-weight:700;background:linear-gradient(135deg,#60a5fa,#a78bfa);-webkit-background-clip:text;-webkit-text-fill-color:transparent">Multi-agente Pro</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:13px;color:#64748b;margin-bottom:20px">Investigador + escritor</div>', unsafe_allow_html=True)
    st.markdown("---")

    st.markdown('<div style="font-size:12px;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px">Agentes</div>', unsafe_allow_html=True)

    agents_info = [
        ("🎯", "Coordinador", "Analiza y planifica", "#AFA9EC"),
        ("🔍", "Investigador", "Busca datos y fuentes", "#5DCAA5"),
        ("✍️", "Escritor", "Redacta documentos", "#F0997B"),
    ]
    for icon, name, desc, color in agents_info:
        st.markdown(f'<div style="background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.06);border-radius:10px;padding:10px 14px;margin:6px 0;display:flex;align-items:center;gap:10px"><span style="font-size:18px">{icon}</span><div><div style="font-size:13px;color:{color};font-weight:500">{name}</div><div style="font-size:11px;color:#64748b">{desc}</div></div></div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div style="font-size:12px;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px">Ejemplos</div>', unsafe_allow_html=True)

    examples = [
        "Investiga sobre IA en México y escribe un artículo",
        "Busca noticias de tecnología y hazme un resumen ejecutivo",
        "Investiga energía solar y escribe un reporte en energia.md",
        "Escribe un email profesional presentando una empresa de software",
    ]
    for ex in examples:
        if st.button(f"💡 {ex}", key=f"ex_{ex[:20]}", use_container_width=True):
            st.session_state.pending_input = ex
            st.rerun()

    st.markdown("---")
    if st.button("🗑️ Limpiar", use_container_width=True):
        st.session_state.history = []
        st.rerun()

    st.markdown('<div style="font-size:11px;color:#334155;text-align:center;margin-top:16px">by Oswaldo Murillo</div>', unsafe_allow_html=True)


# Header
st.markdown('<div class="header-title">Sistema multi-agente</div>', unsafe_allow_html=True)
st.markdown('<div class="header-sub">Un investigador y un escritor colaboran para resolver tu solicitud</div>', unsafe_allow_html=True)

# Mostrar historial
for entry in st.session_state.history:
    if entry["type"] == "user":
        st.markdown(f'<div class="user-msg">{entry["content"]}</div>', unsafe_allow_html=True)
    elif entry["type"] == "agent":
        css_class = entry.get("css", "coordinator")
        st.markdown(f'<div class="agent-card {css_class}"><div class="agent-name"><span class="{entry.get("color_class","coord-color")}">{entry["name"]}</span></div><div class="agent-content">{entry["content"]}</div></div>', unsafe_allow_html=True)
    elif entry["type"] == "tools":
        pills = "".join(f'<span class="tool-pill">🔧 {t}</span>' for t in entry["tools"])
        st.markdown(f'<div style="margin:4px 0">{pills}</div>', unsafe_allow_html=True)
    elif entry["type"] == "arrow":
        st.markdown('<div class="flow-arrow">↓</div>', unsafe_allow_html=True)
    elif entry["type"] == "result":
        st.markdown(f'<div class="result-card">{entry["content"]}</div>', unsafe_allow_html=True)

# Input
solicitud = st.chat_input("Describe tu tarea (ej: Investiga sobre IA y escribe un artículo)")

if "pending_input" in st.session_state:
    solicitud = st.session_state.pending_input
    del st.session_state.pending_input

if solicitud:
    st.session_state.history.append({"type": "user", "content": solicitud})
    st.markdown(f'<div class="user-msg">{solicitud}</div>', unsafe_allow_html=True)

    # Paso 1: Coordinador
    with st.spinner("🎯 Coordinador analizando..."):
        plan = planificar(solicitud)

    plan_text = f"📋 Plan: {plan['plan']}\n📊 {len(plan['pasos'])} paso(s)"
    st.session_state.history.append({"type": "agent", "name": "🎯 Coordinador", "content": plan_text, "css": "coordinator", "color_class": "coord-color"})
    st.markdown(f'<div class="agent-card coordinator"><div class="agent-name"><span class="coord-color">🎯 Coordinador</span></div><div class="agent-content">{plan_text}</div></div>', unsafe_allow_html=True)

    st.session_state.history.append({"type": "arrow"})
    st.markdown('<div class="flow-arrow">↓</div>', unsafe_allow_html=True)

    # Paso 2: Ejecutar agentes
    contexto = ""
    resultado_final = ""

    for i, paso in enumerate(plan["pasos"]):
        agente = paso["agente"].lower()
        tarea = paso["tarea"]

        if agente == "investigador":
            with st.spinner(f"🔍 Investigador trabajando..."):
                resultado, tools = ejecutar_agente(
                    "Investigador", "🔍", PROMPT_INVESTIGADOR,
                    HERRAMIENTAS_INV, FUNCIONES_INV, tarea, contexto)

            if tools:
                st.session_state.history.append({"type": "tools", "tools": tools})
                pills = "".join(f'<span class="tool-pill">🔧 {t}</span>' for t in tools)
                st.markdown(f'<div style="margin:4px 0">{pills}</div>', unsafe_allow_html=True)

            preview = resultado[:300] + "..." if len(resultado) > 300 else resultado
            st.session_state.history.append({"type": "agent", "name": "🔍 Investigador", "content": preview, "css": "researcher", "color_class": "research-color"})
            st.markdown(f'<div class="agent-card researcher"><div class="agent-name"><span class="research-color">🔍 Investigador</span></div><div class="agent-content">{preview}</div></div>', unsafe_allow_html=True)

            contexto += f"\nDATOS DEL INVESTIGADOR:\n{resultado}"

        elif agente == "escritor":
            if i > 0:
                st.session_state.history.append({"type": "arrow"})
                st.markdown('<div class="flow-arrow">↓</div>', unsafe_allow_html=True)

            with st.spinner(f"✍️ Escritor redactando..."):
                resultado, tools = ejecutar_agente(
                    "Escritor", "✍️", PROMPT_ESCRITOR,
                    HERRAMIENTAS_ESC, FUNCIONES_ESC, tarea, contexto)

            if tools:
                st.session_state.history.append({"type": "tools", "tools": tools})
                pills = "".join(f'<span class="tool-pill">🔧 {t}</span>' for t in tools)
                st.markdown(f'<div style="margin:4px 0">{pills}</div>', unsafe_allow_html=True)

            resultado_final = resultado

        if i < len(plan["pasos"]) - 1 and agente == "investigador":
            st.session_state.history.append({"type": "arrow"})
            st.markdown('<div class="flow-arrow">↓</div>', unsafe_allow_html=True)

    # Resultado final
    if not resultado_final:
        resultado_final = contexto

    st.session_state.history.append({"type": "arrow"})
    st.markdown('<div class="flow-arrow">↓</div>', unsafe_allow_html=True)

    st.session_state.history.append({"type": "result", "content": resultado_final})
    st.markdown(f'<div class="result-card">{resultado_final}</div>', unsafe_allow_html=True)

    st.rerun()
