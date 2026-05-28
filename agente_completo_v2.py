"""
=============================================================
🤖 AGENTE PRO v2 - Con Herramientas Reales (Mejorado)
=============================================================

Este agente tiene 6 herramientas reales:
  1. 🔍 Buscar en la web (DuckDuckGo con reintentos + búsqueda de noticias)
  2. 📧 Enviar correos electrónicos (usando SMTP de Gmail)
  3. 📄 Leer PDFs
  4. 📊 Leer archivos Excel/CSV
  5. 🗄️ Consultar base de datos SQLite
  6. 💾 Guardar notas/resultados en archivos

INSTALACIÓN:
------------
Ejecuta este comando para instalar todas las dependencias:

  pip install anthropic python-dotenv duckduckgo-search PyPDF2 openpyxl

CONFIGURACIÓN (.env):
---------------------
  ANTHROPIC_API_KEY=sk-ant-tu-api-key
  EMAIL_ADDRESS=tu-correo@gmail.com
  EMAIL_PASSWORD=tu-contraseña-de-aplicacion

NOTA SOBRE GMAIL:
  Para enviar correos necesitas una "Contraseña de aplicación" de Google.
  Ve a myaccount.google.com → Seguridad → Verificación en 2 pasos (actívala)
  → Contraseñas de aplicaciones → Genera una para "Correo".
  Esa contraseña de 16 caracteres es la que pones en EMAIL_PASSWORD.

=============================================================
"""

import os
import json
import sqlite3
import smtplib
import csv
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from anthropic import Anthropic

load_dotenv()


# =============================================================
# HERRAMIENTA 1: 🔍 BÚSQUEDA WEB (DuckDuckGo - gratis)
# =============================================================

def buscar_en_web(consulta: str, max_resultados: int = 5) -> str:
    """
    Busca información en la web usando DuckDuckGo.
    Incluye reintentos y headers para evitar bloqueos.
    """
    import time

    try:
        from duckduckgo_search import DDGS
    except ImportError:
        return "Error: Instala duckduckgo-search con: pip install duckduckgo-search"

    # Intento 1: búsqueda de texto normal
    for intento in range(3):
        try:
            with DDGS() as ddgs:
                resultados = list(ddgs.text(
                    consulta,
                    max_results=max_resultados,
                    region="wt-wt"
                ))

            if resultados:
                texto = f"Resultados de búsqueda para '{consulta}':\n\n"
                for i, r in enumerate(resultados, 1):
                    texto += f"{i}. **{r['title']}**\n"
                    texto += f"   {r['body']}\n"
                    texto += f"   Fuente: {r['href']}\n\n"
                return texto

        except Exception:
            if intento < 2:
                time.sleep(2)
                continue

    # Intento 2: búsqueda de noticias como alternativa
    try:
        with DDGS() as ddgs:
            resultados = list(ddgs.news(consulta, max_results=max_resultados))

        if resultados:
            texto = f"Noticias encontradas para '{consulta}':\n\n"
            for i, r in enumerate(resultados, 1):
                texto += f"{i}. **{r['title']}**\n"
                texto += f"   {r['body']}\n"
                texto += f"   Fuente: {r['source']}\n"
                texto += f"   Fecha: {r.get('date', 'N/A')}\n\n"
            return texto

    except Exception:
        pass

    return (f"No se pudieron obtener resultados para '{consulta}' después de varios intentos. "
            "DuckDuckGo puede estar limitando las solicitudes temporalmente. "
            "Intenta de nuevo en unos minutos o reformula la búsqueda.")


# =============================================================
# HERRAMIENTA 2: 📧 ENVIAR CORREO ELECTRÓNICO (Gmail SMTP)
# =============================================================

def enviar_correo(destinatario: str, asunto: str, cuerpo: str) -> str:
    """
    Envía un correo electrónico usando Gmail SMTP.
    Requiere EMAIL_ADDRESS y EMAIL_PASSWORD en el archivo .env
    """
    email_address = os.getenv("EMAIL_ADDRESS")
    email_password = os.getenv("EMAIL_PASSWORD")

    if not email_address or not email_password:
        return ("Error: Configura EMAIL_ADDRESS y EMAIL_PASSWORD en tu archivo .env. "
                "Necesitas una contraseña de aplicación de Google "
                "(ve a myaccount.google.com → Seguridad → Contraseñas de aplicaciones).")

    try:
        mensaje = MIMEMultipart()
        mensaje["From"] = email_address
        mensaje["To"] = destinatario
        mensaje["Subject"] = asunto
        mensaje.attach(MIMEText(cuerpo, "plain"))

        with smtplib.SMTP("smtp.gmail.com", 587) as servidor:
            servidor.starttls()
            servidor.login(email_address, email_password)
            servidor.send_message(mensaje)

        return f"Correo enviado exitosamente a {destinatario} con asunto '{asunto}'."

    except smtplib.SMTPAuthenticationError:
        return ("Error de autenticación. Verifica tu EMAIL_PASSWORD. "
                "Recuerda que necesitas una 'Contraseña de aplicación' de Google, "
                "no tu contraseña normal de Gmail.")
    except Exception as e:
        return f"Error al enviar correo: {str(e)}"


# =============================================================
# HERRAMIENTA 3: 📄 LEER ARCHIVOS PDF
# =============================================================

def leer_pdf(ruta_archivo: str) -> str:
    """
    Lee y extrae el texto de un archivo PDF.
    """
    try:
        from PyPDF2 import PdfReader

        ruta = Path(ruta_archivo)
        if not ruta.exists():
            return f"Error: No se encontró el archivo '{ruta_archivo}'."

        if not ruta.suffix.lower() == ".pdf":
            return f"Error: El archivo '{ruta_archivo}' no es un PDF."

        reader = PdfReader(str(ruta))
        texto_completo = ""

        for i, pagina in enumerate(reader.pages, 1):
            texto = pagina.extract_text()
            if texto:
                texto_completo += f"\n--- Página {i} ---\n{texto}"

        if not texto_completo.strip():
            return "El PDF no contiene texto extraíble (puede ser un PDF escaneado/imagen)."

        # Limitar a 4000 caracteres para no exceder el contexto
        if len(texto_completo) > 4000:
            texto_completo = texto_completo[:4000] + "\n\n[... texto truncado por longitud ...]"

        return f"Contenido del PDF '{ruta.name}' ({len(reader.pages)} páginas):\n{texto_completo}"

    except ImportError:
        return "Error: Instala PyPDF2 con: pip install PyPDF2"
    except Exception as e:
        return f"Error al leer PDF: {str(e)}"


# =============================================================
# HERRAMIENTA 4: 📊 LEER ARCHIVOS EXCEL Y CSV
# =============================================================

def leer_excel_csv(ruta_archivo: str, max_filas: int = 20) -> str:
    """
    Lee archivos Excel (.xlsx, .xls) y CSV (.csv).
    Muestra las primeras filas y un resumen.
    """
    try:
        ruta = Path(ruta_archivo)
        if not ruta.exists():
            return f"Error: No se encontró el archivo '{ruta_archivo}'."

        extension = ruta.suffix.lower()

        if extension == ".csv":
            filas = []
            with open(ruta, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                for i, fila in enumerate(reader):
                    if i >= max_filas + 1:  # +1 por el encabezado
                        break
                    filas.append(fila)

            if not filas:
                return "El archivo CSV está vacío."

            encabezados = filas[0]
            datos = filas[1:]
            total_filas = sum(1 for _ in open(ruta, "r", encoding="utf-8-sig")) - 1

        elif extension in [".xlsx", ".xls"]:
            try:
                from openpyxl import load_workbook
            except ImportError:
                return "Error: Instala openpyxl con: pip install openpyxl"

            wb = load_workbook(str(ruta), read_only=True)
            ws = wb.active

            filas = []
            for i, fila in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_filas + 1:
                    break
                filas.append([str(c) if c is not None else "" for c in fila])

            wb.close()

            if not filas:
                return "El archivo Excel está vacío."

            encabezados = filas[0]
            datos = filas[1:]
            total_filas = ws.max_row - 1 if ws.max_row else 0

        else:
            return f"Error: Formato no soportado '{extension}'. Usa .csv, .xlsx o .xls"

        # Formatear como tabla
        resultado = f"Archivo: {ruta.name}\n"
        resultado += f"Columnas: {', '.join(encabezados)}\n"
        resultado += f"Total de filas: {total_filas}\n"
        resultado += f"Mostrando primeras {min(len(datos), max_filas)} filas:\n\n"

        # Encabezados
        resultado += " | ".join(encabezados) + "\n"
        resultado += "-" * (len(resultado.split("\n")[-2])) + "\n"

        # Datos
        for fila in datos:
            resultado += " | ".join(fila) + "\n"

        if total_filas > max_filas:
            resultado += f"\n[... {total_filas - max_filas} filas más no mostradas ...]"

        return resultado

    except Exception as e:
        return f"Error al leer archivo: {str(e)}"


# =============================================================
# HERRAMIENTA 5: 🗄️ BASE DE DATOS SQLite
# =============================================================

DB_PATH = "agente_datos.db"


def inicializar_db():
    """Crea la base de datos y tabla de ejemplo si no existen."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Tabla de ejemplo: contactos
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS contactos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            email TEXT,
            telefono TEXT,
            notas TEXT,
            creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Tabla de ejemplo: tareas
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tareas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT NOT NULL,
            descripcion TEXT,
            estado TEXT DEFAULT 'pendiente',
            prioridad TEXT DEFAULT 'media',
            creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Insertar datos de ejemplo si las tablas están vacías
    cursor.execute("SELECT COUNT(*) FROM contactos")
    if cursor.fetchone()[0] == 0:
        contactos_ejemplo = [
            ("Ana García", "ana@ejemplo.com", "33-1234-5678", "Cliente principal"),
            ("Carlos López", "carlos@ejemplo.com", "81-9876-5432", "Proveedor"),
            ("María Rodríguez", "maria@ejemplo.com", "55-5555-1234", "Socia de proyecto"),
        ]
        cursor.executemany(
            "INSERT INTO contactos (nombre, email, telefono, notas) VALUES (?, ?, ?, ?)",
            contactos_ejemplo
        )

    cursor.execute("SELECT COUNT(*) FROM tareas")
    if cursor.fetchone()[0] == 0:
        tareas_ejemplo = [
            ("Revisar propuesta", "Revisar la propuesta del cliente", "pendiente", "alta"),
            ("Enviar reporte", "Reporte mensual de ventas", "pendiente", "media"),
            ("Llamar a proveedor", "Negociar precios Q3", "completada", "baja"),
        ]
        cursor.executemany(
            "INSERT INTO tareas (titulo, descripcion, estado, prioridad) VALUES (?, ?, ?, ?)",
            tareas_ejemplo
        )

    conn.commit()
    conn.close()


def consultar_base_datos(consulta_sql: str) -> str:
    """
    Ejecuta consultas SQL en la base de datos SQLite.
    Solo permite SELECT, INSERT y UPDATE (no DELETE ni DROP por seguridad).
    """
    consulta_upper = consulta_sql.strip().upper()

    # Seguridad: solo permitir operaciones seguras
    if any(palabra in consulta_upper for palabra in ["DROP", "DELETE", "ALTER", "TRUNCATE"]):
        return "Error: Por seguridad, las operaciones DROP, DELETE, ALTER y TRUNCATE no están permitidas."

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute(consulta_sql)

        if consulta_upper.startswith("SELECT"):
            columnas = [desc[0] for desc in cursor.description] if cursor.description else []
            filas = cursor.fetchall()

            if not filas:
                resultado = "La consulta no devolvió resultados."
            else:
                resultado = f"Columnas: {', '.join(columnas)}\n"
                resultado += f"Resultados ({len(filas)} filas):\n\n"
                for fila in filas:
                    resultado += " | ".join(str(v) for v in fila) + "\n"
        else:
            conn.commit()
            resultado = f"Consulta ejecutada exitosamente. Filas afectadas: {cursor.rowcount}"

        conn.close()
        return resultado

    except sqlite3.Error as e:
        return f"Error SQL: {str(e)}"
    except Exception as e:
        return f"Error: {str(e)}"


# =============================================================
# HERRAMIENTA 6: 💾 GUARDAR NOTAS/RESULTADOS EN ARCHIVO
# =============================================================

def guardar_archivo(nombre_archivo: str, contenido: str) -> str:
    """
    Guarda texto en un archivo .txt o .md
    """
    try:
        ruta = Path(nombre_archivo)

        # Solo permitir extensiones seguras
        extensiones_permitidas = [".txt", ".md", ".csv", ".json"]
        if ruta.suffix.lower() not in extensiones_permitidas:
            return f"Error: Solo se permiten archivos {', '.join(extensiones_permitidas)}"

        with open(ruta, "w", encoding="utf-8") as f:
            f.write(contenido)

        return f"Archivo '{nombre_archivo}' guardado exitosamente ({len(contenido)} caracteres)."

    except Exception as e:
        return f"Error al guardar archivo: {str(e)}"


# =============================================================
# HERRAMIENTA 7: 🌐 TRADUCTOR DE IDIOMAS
# =============================================================

def traducir_texto(texto: str, idioma_destino: str, idioma_origen: str = "auto") -> str:
    """
    Traduce texto entre idiomas usando Claude como motor de traducción.
    No necesita API externa - usa el mismo modelo del agente.
    """
    try:
        client = Anthropic()

        prompt = f"""Traduce el siguiente texto al {idioma_destino}.
Si el idioma de origen es "auto", detecta automáticamente el idioma.
Idioma de origen: {idioma_origen}

REGLAS:
- Devuelve SOLO la traducción, sin explicaciones ni comentarios
- Mantén el formato original (saltos de línea, viñetas, etc.)
- Si hay nombres propios, no los traduzcas

Texto a traducir:
{texto}"""

        respuesta = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}]
        )

        traduccion = ""
        for bloque in respuesta.content:
            if hasattr(bloque, "text"):
                traduccion += bloque.text

        idioma_info = f" (detectado automáticamente)" if idioma_origen == "auto" else ""
        return (
            f"Traducción al {idioma_destino}{idioma_info}:\n\n"
            f"{traduccion}\n\n"
            f"--- Texto original ({len(texto)} caracteres) → "
            f"Traducción ({len(traduccion)} caracteres) ---"
        )

    except Exception as e:
        return f"Error al traducir: {str(e)}"


# =============================================================
# HERRAMIENTA 8: 📈 GENERAR GRÁFICAS
# =============================================================

def generar_grafica(
    ruta_csv: str,
    tipo_grafica: str = "barras",
    columna_x: str = "",
    columna_y: str = "",
    titulo: str = "Gráfica",
    nombre_archivo: str = "grafica.png"
) -> str:
    """
    Genera gráficas a partir de datos CSV o Excel.
    Tipos: barras, lineas, pastel, dispersión, barras_horizontales.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")  # Backend sin GUI (para servidores/scripts)
        import matplotlib.pyplot as plt
        import matplotlib.ticker as ticker
    except ImportError:
        return "Error: Instala matplotlib con: pip install matplotlib"

    try:
        ruta = Path(ruta_csv)
        if not ruta.exists():
            return f"Error: No se encontró el archivo '{ruta_csv}'."

        # Leer datos según extensión
        extension = ruta.suffix.lower()
        datos = []
        encabezados = []

        if extension == ".csv":
            with open(ruta, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                encabezados = next(reader)
                for fila in reader:
                    datos.append(fila)

        elif extension in [".xlsx", ".xls"]:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(ruta), read_only=True)
                ws = wb.active
                filas = list(ws.iter_rows(values_only=True))
                wb.close()
                if filas:
                    encabezados = [str(c) if c else f"Col{i}" for i, c in enumerate(filas[0])]
                    datos = [[str(c) if c else "" for c in fila] for fila in filas[1:]]
            except ImportError:
                return "Error: Instala openpyxl con: pip install openpyxl"
        else:
            return f"Error: Formato no soportado '{extension}'. Usa .csv o .xlsx"

        if not datos or not encabezados:
            return "Error: El archivo está vacío o no tiene datos válidos."

        # Auto-detectar columnas si no se especificaron
        if not columna_x:
            columna_x = encabezados[0]
        if not columna_y:
            # Buscar la primera columna numérica
            for i, header in enumerate(encabezados[1:], 1):
                try:
                    float(datos[0][i].replace(",", "").replace("$", ""))
                    columna_y = header
                    break
                except (ValueError, IndexError):
                    continue
            if not columna_y:
                columna_y = encabezados[1] if len(encabezados) > 1 else encabezados[0]

        # Encontrar índices de columnas
        col_x_idx = None
        col_y_idx = None
        for i, h in enumerate(encabezados):
            if h.strip().lower() == columna_x.strip().lower():
                col_x_idx = i
            if h.strip().lower() == columna_y.strip().lower():
                col_y_idx = i

        if col_x_idx is None:
            return f"Error: Columna '{columna_x}' no encontrada. Columnas disponibles: {', '.join(encabezados)}"
        if col_y_idx is None:
            return f"Error: Columna '{columna_y}' no encontrada. Columnas disponibles: {', '.join(encabezados)}"

        # Extraer valores
        etiquetas_x = []
        valores_y = []

        # Agrupar si hay valores repetidos en X
        agrupado = {}
        for fila in datos:
            try:
                x_val = fila[col_x_idx].strip()
                y_val = fila[col_y_idx].strip().replace(",", "").replace("$", "")
                y_num = float(y_val)

                if x_val in agrupado:
                    agrupado[x_val] += y_num
                else:
                    agrupado[x_val] = y_num
            except (ValueError, IndexError):
                continue

        if not agrupado:
            return "Error: No se pudieron extraer datos numéricos de la columna Y."

        etiquetas_x = list(agrupado.keys())
        valores_y = list(agrupado.values())

        # Configurar estilo de la gráfica
        plt.style.use("seaborn-v0_8-whitegrid")
        fig, ax = plt.subplots(figsize=(10, 6))

        # Colores profesionales
        colores = ["#378ADD", "#1D9E75", "#D85A30", "#D4537E", "#BA7517",
                    "#534AB7", "#639922", "#E24B4A", "#888780"]

        tipo = tipo_grafica.lower().strip()

        if tipo in ["barras", "bar"]:
            barras = ax.bar(etiquetas_x, valores_y, color=colores[:len(etiquetas_x)],
                           edgecolor="white", linewidth=0.5)
            # Agregar valores encima de cada barra
            for barra, valor in zip(barras, valores_y):
                ax.text(barra.get_x() + barra.get_width()/2., barra.get_height(),
                       f'{valor:,.0f}', ha='center', va='bottom', fontsize=9, fontweight='bold')

        elif tipo in ["lineas", "linea", "line"]:
            ax.plot(etiquetas_x, valores_y, color=colores[0], marker="o",
                   linewidth=2, markersize=8, markerfacecolor="white",
                   markeredgecolor=colores[0], markeredgewidth=2)
            # Agregar valores en cada punto
            for i, (x, y) in enumerate(zip(etiquetas_x, valores_y)):
                ax.annotate(f'{y:,.0f}', (x, y), textcoords="offset points",
                           xytext=(0, 12), ha='center', fontsize=9, fontweight='bold')

        elif tipo in ["pastel", "pie"]:
            ax.pie(valores_y, labels=etiquetas_x, colors=colores[:len(etiquetas_x)],
                  autopct='%1.1f%%', startangle=90, textprops={'fontsize': 11})
            ax.axis('equal')

        elif tipo in ["barras_horizontales", "horizontal", "barh"]:
            barras = ax.barh(etiquetas_x, valores_y, color=colores[:len(etiquetas_x)],
                            edgecolor="white", linewidth=0.5)
            for barra, valor in zip(barras, valores_y):
                ax.text(barra.get_width() + max(valores_y)*0.01, barra.get_y() + barra.get_height()/2.,
                       f'{valor:,.0f}', ha='left', va='center', fontsize=9, fontweight='bold')

        elif tipo in ["dispersión", "dispersion", "scatter"]:
            ax.scatter(range(len(valores_y)), valores_y, c=colores[0], s=100,
                      edgecolors="white", linewidth=1, zorder=5)
            ax.set_xticks(range(len(etiquetas_x)))
            ax.set_xticklabels(etiquetas_x)

        else:
            return f"Error: Tipo de gráfica '{tipo_grafica}' no soportado. Usa: barras, lineas, pastel, barras_horizontales, dispersión"

        # Formato general
        ax.set_title(titulo, fontsize=16, fontweight="bold", pad=15)
        if tipo != "pastel":
            ax.set_xlabel(columna_x, fontsize=12, labelpad=10)
            ax.set_ylabel(columna_y, fontsize=12, labelpad=10)
            ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, p: f'{x:,.0f}'))
            plt.xticks(rotation=45 if len(etiquetas_x) > 5 else 0, ha='right')

        plt.tight_layout()

        # Guardar
        ruta_salida = Path(nombre_archivo)
        if ruta_salida.suffix.lower() not in [".png", ".jpg", ".pdf", ".svg"]:
            nombre_archivo = nombre_archivo + ".png"

        plt.savefig(nombre_archivo, dpi=150, bbox_inches='tight',
                   facecolor='white', edgecolor='none')
        plt.close()

        return (
            f"Gráfica generada exitosamente:\n"
            f"  Archivo: {nombre_archivo}\n"
            f"  Tipo: {tipo_grafica}\n"
            f"  Datos: {columna_x} vs {columna_y}\n"
            f"  Puntos de datos: {len(etiquetas_x)}\n"
            f"  Valores: {', '.join(f'{e}={v:,.0f}' for e, v in zip(etiquetas_x, valores_y))}\n\n"
            f"La gráfica se guardó en la carpeta actual. Ábrela con cualquier visor de imágenes."
        )

    except Exception as e:
        return f"Error al generar gráfica: {str(e)}"


# =============================================================
# DEFINICIÓN DE HERRAMIENTAS PARA CLAUDE
# =============================================================

HERRAMIENTAS = [
    {
        "name": "buscar_en_web",
        "description": (
            "Busca información actualizada en internet usando DuckDuckGo. "
            "Úsala para cualquier pregunta sobre eventos recientes, noticias, "
            "datos actualizados, o cualquier tema que requiera información en tiempo real."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "consulta": {
                    "type": "string",
                    "description": "Lo que quieres buscar en internet"
                },
                "max_resultados": {
                    "type": "integer",
                    "description": "Número máximo de resultados (1-10, por defecto 5)",
                    "default": 5
                }
            },
            "required": ["consulta"]
        }
    },
    {
        "name": "enviar_correo",
        "description": (
            "Envía un correo electrónico a una dirección específica. "
            "Úsala cuando el usuario pida enviar un email, notificación o mensaje."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "destinatario": {
                    "type": "string",
                    "description": "Dirección de correo del destinatario"
                },
                "asunto": {
                    "type": "string",
                    "description": "Asunto del correo"
                },
                "cuerpo": {
                    "type": "string",
                    "description": "Contenido del correo"
                }
            },
            "required": ["destinatario", "asunto", "cuerpo"]
        }
    },
    {
        "name": "leer_pdf",
        "description": (
            "Lee y extrae el texto de un archivo PDF. "
            "Úsala cuando el usuario quiera leer, analizar o resumir un PDF. "
            "Necesita la ruta completa al archivo."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "ruta_archivo": {
                    "type": "string",
                    "description": "Ruta completa al archivo PDF (ej: C:/Users/docs/archivo.pdf)"
                }
            },
            "required": ["ruta_archivo"]
        }
    },
    {
        "name": "leer_excel_csv",
        "description": (
            "Lee archivos Excel (.xlsx) o CSV (.csv) y muestra su contenido. "
            "Úsala cuando el usuario quiera ver, analizar o consultar datos de una hoja de cálculo."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "ruta_archivo": {
                    "type": "string",
                    "description": "Ruta completa al archivo Excel o CSV"
                },
                "max_filas": {
                    "type": "integer",
                    "description": "Máximo de filas a mostrar (por defecto 20)",
                    "default": 20
                }
            },
            "required": ["ruta_archivo"]
        }
    },
    {
        "name": "consultar_base_datos",
        "description": (
            "Ejecuta consultas SQL en la base de datos. "
            "La base de datos tiene 2 tablas: "
            "'contactos' (id, nombre, email, telefono, notas, creado_en) y "
            "'tareas' (id, titulo, descripcion, estado, prioridad, creado_en). "
            "Puede hacer SELECT para consultar e INSERT/UPDATE para modificar datos."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "consulta_sql": {
                    "type": "string",
                    "description": "Consulta SQL a ejecutar (SELECT, INSERT o UPDATE)"
                }
            },
            "required": ["consulta_sql"]
        }
    },
    {
        "name": "guardar_archivo",
        "description": (
            "Guarda texto en un archivo. Útil para guardar resultados de búsquedas, "
            "resúmenes, reportes o cualquier información que el usuario quiera conservar."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "nombre_archivo": {
                    "type": "string",
                    "description": "Nombre del archivo (con extensión .txt, .md, .csv o .json)"
                },
                "contenido": {
                    "type": "string",
                    "description": "Texto a guardar en el archivo"
                }
            },
            "required": ["nombre_archivo", "contenido"]
        }
    },
    {
        "name": "traducir_texto",
        "description": (
            "Traduce texto entre idiomas. Soporta español, inglés, francés, "
            "portugués, alemán, italiano, japonés, chino, coreano y muchos más. "
            "Úsala cuando el usuario pida traducir algo, o cuando necesite "
            "entender texto en otro idioma."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "texto": {
                    "type": "string",
                    "description": "El texto que se quiere traducir"
                },
                "idioma_destino": {
                    "type": "string",
                    "description": "Idioma al que se quiere traducir (ej: inglés, francés, japonés)"
                },
                "idioma_origen": {
                    "type": "string",
                    "description": "Idioma del texto original. Usa 'auto' para detectar automáticamente",
                    "default": "auto"
                }
            },
            "required": ["texto", "idioma_destino"]
        }
    },
    {
        "name": "generar_grafica",
        "description": (
            "Genera gráficas profesionales a partir de archivos CSV o Excel. "
            "Tipos disponibles: barras, lineas, pastel, barras_horizontales, dispersión. "
            "Úsala cuando el usuario pida crear una gráfica, visualizar datos, "
            "hacer un chart o mostrar tendencias de un archivo de datos."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "ruta_csv": {
                    "type": "string",
                    "description": "Ruta completa al archivo CSV o Excel con los datos"
                },
                "tipo_grafica": {
                    "type": "string",
                    "description": "Tipo de gráfica: barras, lineas, pastel, barras_horizontales, dispersión",
                    "default": "barras"
                },
                "columna_x": {
                    "type": "string",
                    "description": "Nombre de la columna para el eje X (horizontal). Si se omite, usa la primera columna",
                    "default": ""
                },
                "columna_y": {
                    "type": "string",
                    "description": "Nombre de la columna para el eje Y (valores). Si se omite, usa la primera columna numérica",
                    "default": ""
                },
                "titulo": {
                    "type": "string",
                    "description": "Título de la gráfica",
                    "default": "Gráfica"
                },
                "nombre_archivo": {
                    "type": "string",
                    "description": "Nombre del archivo de imagen a guardar (ej: ventas.png)",
                    "default": "grafica.png"
                }
            },
            "required": ["ruta_csv"]
        }
    }
]


# Mapeo de funciones
FUNCIONES_DISPONIBLES = {
    "buscar_en_web": lambda args: buscar_en_web(
        args["consulta"], args.get("max_resultados", 5)
    ),
    "enviar_correo": lambda args: enviar_correo(
        args["destinatario"], args["asunto"], args["cuerpo"]
    ),
    "leer_pdf": lambda args: leer_pdf(args["ruta_archivo"]),
    "leer_excel_csv": lambda args: leer_excel_csv(
        args["ruta_archivo"], args.get("max_filas", 20)
    ),
    "consultar_base_datos": lambda args: consultar_base_datos(args["consulta_sql"]),
    "guardar_archivo": lambda args: guardar_archivo(
        args["nombre_archivo"], args["contenido"]
    ),
    "traducir_texto": lambda args: traducir_texto(
        args["texto"], args["idioma_destino"], args.get("idioma_origen", "auto")
    ),
    "generar_grafica": lambda args: generar_grafica(
        args["ruta_csv"],
        args.get("tipo_grafica", "barras"),
        args.get("columna_x", ""),
        args.get("columna_y", ""),
        args.get("titulo", "Gráfica"),
        args.get("nombre_archivo", "grafica.png")
    ),
}


# =============================================================
# SYSTEM PROMPT DEL AGENTE
# =============================================================

SYSTEM_PROMPT = """
Eres un asistente personal inteligente y completo que habla español.
Tu nombre es "Agente Pro" y tu objetivo es ayudar al usuario con
cualquier tarea usando las herramientas disponibles.

## TUS HERRAMIENTAS:

1. **buscar_en_web** - Buscar información actualizada en internet
2. **enviar_correo** - Enviar correos electrónicos
3. **leer_pdf** - Leer y extraer texto de archivos PDF
4. **leer_excel_csv** - Leer hojas de cálculo Excel y archivos CSV
5. **consultar_base_datos** - Consultar y modificar la base de datos
   (tablas: contactos y tareas)
6. **guardar_archivo** - Guardar resultados en archivos de texto
7. **traducir_texto** - Traducir texto entre idiomas
8. **generar_grafica** - Crear gráficas a partir de datos CSV/Excel

## REGLAS:

1. Siempre responde en español de forma amigable y profesional
2. Usa las herramientas cuando sea necesario - no inventes datos
3. Para búsquedas web, resume los resultados de forma clara
4. Antes de enviar un correo, confirma con el usuario el contenido
5. Si no puedes hacer algo, explica por qué y sugiere alternativas
6. Cuando uses la base de datos, explica qué consulta estás haciendo
7. Si el usuario te pide guardar algo, usa un nombre de archivo descriptivo
8. Puedes combinar múltiples herramientas para tareas complejas

## EJEMPLOS DE TAREAS QUE PUEDES HACER:

- "Busca noticias sobre IA y guárdalas en un archivo"
  → Usa buscar_en_web + guardar_archivo
- "Lee mi PDF y envíame un resumen por correo"
  → Usa leer_pdf + enviar_correo
- "Agrega un nuevo contacto: Juan Pérez, juan@mail.com"
  → Usa consultar_base_datos con INSERT
- "¿Qué tareas tengo pendientes?"
  → Usa consultar_base_datos con SELECT
- "Traduce esto al inglés: Buenos días, ¿cómo estás?"
  → Usa traducir_texto
- "Busca noticias sobre IA, tradúcelas al inglés y guárdalas"
  → Usa buscar_en_web + traducir_texto + guardar_archivo
- "Genera una gráfica de barras con los datos de ventas.csv"
  → Usa generar_grafica
- "Lee ventas.csv, dime el resumen y genera una gráfica de pastel"
  → Usa leer_excel_csv + generar_grafica
"""


# =============================================================
# LOOP PRINCIPAL DEL AGENTE
# =============================================================

def ejecutar_herramienta(nombre: str, argumentos: dict) -> str:
    """Ejecuta una herramienta por su nombre."""
    if nombre in FUNCIONES_DISPONIBLES:
        return FUNCIONES_DISPONIBLES[nombre](argumentos)
    return f"Herramienta '{nombre}' no encontrada."


def agente_completo():
    """Loop principal del agente con todas las herramientas reales."""

    # Inicializar base de datos
    inicializar_db()

    client = Anthropic()
    mensajes = []

    print("")
    print("=" * 56)
    print("  🤖 AGENTE PRO v2 - Tu Asistente Personal Inteligente")
    print("=" * 56)
    print("")
    print("  Herramientas disponibles:")
    print("  🔍 Búsqueda web      📧 Enviar correos")
    print("  📄 Leer PDFs          📊 Leer Excel/CSV")
    print("  🗄️  Base de datos      💾 Guardar archivos")
    print("  🌐 Traductor           📈 Generar gráficas")
    print("")
    print("  Ejemplos de lo que puedo hacer:")
    print('  • "Busca noticias sobre IA"')
    print('  • "¿Qué contactos tengo en la base de datos?"')
    print('  • "Lee el archivo reporte.pdf"')
    print('  • "Traduce al inglés: Hola, ¿cómo estás?"')
    print('  • "Busca info sobre Python y guárdala en un archivo"')
    print("")
    print('  Escribe "salir" para terminar.')
    print("=" * 56)

    while True:
        pregunta = input("\n🧑 Tú: ").strip()
        if pregunta.lower() in ["salir", "exit", "quit"]:
            print("\n👋 ¡Hasta luego! Tus datos están guardados en la base de datos.")
            break
        if not pregunta:
            continue

        mensajes.append({"role": "user", "content": pregunta})

        # Loop del agente: permite múltiples llamadas a herramientas
        while True:
            try:
                respuesta = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=2048,
                    system=SYSTEM_PROMPT,
                    tools=HERRAMIENTAS,
                    messages=mensajes,
                )
            except Exception as e:
                print(f"\n❌ Error al conectar con Claude: {str(e)}")
                print("   Verifica tu API key y conexión a internet.")
                # Remover el último mensaje para poder reintentar
                mensajes.pop()
                break

            if respuesta.stop_reason == "tool_use":
                mensajes.append({
                    "role": "assistant",
                    "content": respuesta.content
                })

                resultados_herramientas = []
                for bloque in respuesta.content:
                    if bloque.type == "tool_use":
                        nombre = bloque.name
                        args = bloque.input
                        print(f"  🔧 Usando: {nombre}")

                        resultado = ejecutar_herramienta(nombre, args)
                        print(f"  ✅ Listo")

                        resultados_herramientas.append({
                            "type": "tool_result",
                            "tool_use_id": bloque.id,
                            "content": resultado,
                        })

                mensajes.append({
                    "role": "user",
                    "content": resultados_herramientas
                })

            else:
                respuesta_final = ""
                for bloque in respuesta.content:
                    if hasattr(bloque, "text"):
                        respuesta_final += bloque.text

                print(f"\n🤖 Agente Pro: {respuesta_final}")

                mensajes.append({
                    "role": "assistant",
                    "content": respuesta.content
                })
                break


# =============================================================
# PUNTO DE ENTRADA
# =============================================================

if __name__ == "__main__":
    agente_completo()
